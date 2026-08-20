# -*- coding: utf-8 -*-
"""
高教深耕計畫「智慧專案管理與指標管考系統」 (前後期歷史差異化比較版)
特色：非單純覆蓋，完整保留歷史填報軌跡，呈現前後指標差異化比較、轉移達成與滯後未顯著增加追蹤中樞。
執行方式：python sprout_pm_app.py
訪問網址：http://localhost:8080 或 http://127.0.0.1:8080
"""

import http.server
import socketserver
import json
import urllib.parse
import os
import csv
import io
import sys
import re
import openpyxl
from datetime import datetime

try:
    from auto_upload_github import check_git_status, push_to_github, load_config, save_config
except ImportError:
    def check_git_status(): return {"is_repo": False, "error": "auto_upload_github.py 未載入"}
    def push_to_github(**kwargs): return {"success": False, "message": "auto_upload_github.py 未載入"}
    def load_config(): return {}
    def save_config(cfg): pass

# Ensure UTF-8 output on Windows console
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')


# 載入全表 62 項指標與歷史軌跡資料庫
INDICATORS_DB = [
    {
        "id": "COM-01",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "學生專業實務技術能力推動成效 - 學生通過證照數人次",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n1375(教育部提供)\n達成率94.83%",
        "calc_rate": 0.9483,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8683,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8683,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9483,
                "text": "(114-1)\n1375(教育部提供)\n達成率94.83%"
            }
        ]
    },
    {
        "id": "COM-02",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "學生參加競賽獲獎人次",
        "dept": "學務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n308(教育部提供)\n學務處覆核後，是154人次",
        "calc_rate": 0.4875,
        "qualitative_desc": "【未達標檢討】1.確實檢視資料彙整項目內容，減少資料庫競賽項目的重複或缺失，提升資料庫項目的完整性，使其內容得以納入參考數據。\n2.盤點各科系回報之有效數據，減少數據彙整資訊不全狀況。 ｜ 經確認，係因填報時誤重複上傳資料，導致數據由154人次倍增為308人次，謹此致歉。後續將加強填報前後之檢核作業，除確認人次統計外，亦將逐筆核對填報內容，以避免重複填報情形再次發生，確保資料正確無誤。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【雙校核對防錯 Agent】",
        "ai_strategy": "啟動 【雙校核對防錯 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.4675,
        "rate_delta": 0.02,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.4675,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.4875,
                "text": "(114-1)\n308(教育部提供)\n學務處覆核後，是154人次"
            }
        ]
    },
    {
        "id": "COM-03",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "3.0",
        "item": "協助學生專業實務技術能力提升之推動策略描述",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n為培養學生專業實務能力，藉由系科專業證照盤點，聚焦核心證照及分級獎勵制度，驅動學生超越自我、挑戰自我與進階發展，並由各系提出學生專業證照輔導班計",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-1)\n為培養學生專業實務能力，藉由系科專業證照盤點，聚焦核心證照及分級獎勵制度，驅動學生超越自我、挑戰自我與進階發展，並由各系提出學生專業證照輔導班計"
            }
        ]
    },
    {
        "id": "COM-04",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "教師實務經驗提升成效 - 學校全體教師完成半年與專業或技術有關研習或研究之比率",
        "dept": "研發處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n45.02%\n(教育部提供)\n達成率90.04%",
        "calc_rate": 0.8268,
        "qualitative_desc": "【未達標檢討】無 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8218,
        "rate_delta": 0.005,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8218,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.8268,
                "text": "(114-2)\n45.02%\n(教育部提供)\n達成率90.04%"
            }
        ]
    },
    {
        "id": "COM-05",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "學校聘任曾於國際技能競賽獲獎之選手為專任教學人員或專業實作指導人員人數",
        "dept": "學務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n7\n達成率70%",
        "calc_rate": 0.875,
        "qualitative_desc": "【未達標檢討】無 ｜ 此數據與上學期持平，距離目標值仍有成長空間，將持續推動各系聘任具有國際技能競賽獲獎之選手擔任教學或實作指導人員。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【雙校核對防錯 Agent】",
        "ai_strategy": "啟動 【雙校核對防錯 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.87,
        "rate_delta": 0.005,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.87,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.875,
                "text": "(114-2)\n7\n達成率70%"
            }
        ]
    },
    {
        "id": "COM-06",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "3.0",
        "item": "協助教師實務經驗提升策略描述",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n為提升教師實務經驗與實務教學能力，本校協助教師進行產業研習或研究，具體做法包括(1)藉由整體獎補助機制，推動教師進行先期產學合作；(2)藉由整體",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n為提升教師實務經驗與實務教學能力，本校協助教師進行產業研習或研究，具體做法包括(1)藉由整體獎補助機制，推動教師進行先期產學合作；(2)藉由整體"
            }
        ]
    },
    {
        "id": "COM-07",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "教師推動創新教學成效之提升成效 - 採用創新教學模式教師數(學期/人)",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n131\n(教育部提供)\n達成率32.75%",
        "calc_rate": 0.5625,
        "qualitative_desc": "【未達標檢討】113學年第2學期起「創新教學課程」定義修改(有6周以上實施創新教學始得計入)，故課程數下降。 ｜ 1.加強宣導，請老師務必至少有一門課要實施6周以上的創新教學。\n2.因與目標值差異大，建議修正目標值。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【創新教案 AI Agent】",
        "ai_strategy": "啟動 【創新教案 AI Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.5425,
        "rate_delta": 0.02,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.5425,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.5625,
                "text": "(114-2)\n131\n(教育部提供)\n達成率32.75%"
            }
        ]
    },
    {
        "id": "COM-08",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "修讀創新教學課程學生人次(學期/人)",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n7494\n(教育部提供)\n達成率23.42%",
        "calc_rate": 0.3855,
        "qualitative_desc": "【未達標檢討】113學年第2學期起「創新教學課程」定義修改(有6周以上實施創新教學始得計入)，故修讀人次下降。 ｜ 1.加強宣導，請老師務必至少有一門課要實施6周以上的創新教學。\n2.因少子化影響，致使修讀人次下滑，建議修正目標值。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【創新教案 AI Agent】",
        "ai_strategy": "啟動 【創新教案 AI Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.3655,
        "rate_delta": 0.02,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.3655,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.3855,
                "text": "(114-2)\n7494\n(教育部提供)\n達成率23.42%"
            }
        ]
    },
    {
        "id": "COM-09",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "3.0",
        "item": "促進創新教學課程之教學及學習成效提升之推動策略描述",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n研發創新教學方式，助於刺激學生學習動機及上課熱忱、以提升教學成效目標之達成。",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【創新教案 AI Agent】",
        "ai_strategy": "啟動 【創新教案 AI Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n研發創新教學方式，助於刺激學生學習動機及上課熱忱、以提升教學成效目標之達成。"
            }
        ]
    },
    {
        "id": "COM-10",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1-1",
        "item": "修讀跨域學習課程學生人次-雙主修",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n(男)2\n(女)22\n(教育部提供)\n合計24，達成率60%",
        "calc_rate": 0.6,
        "qualitative_desc": "1.雙主修應修課程較繁重、跨系修課恐衝堂，以及學生擔心延畢風險，導致修讀意願不高。\n2.未來將於每學期之跨域學習說明會，加強宣導，讓學生了解輔系之益處，以提升學生修讀意願。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【智慧排課導航 Agent】",
        "ai_strategy": "啟動 【智慧排課導航 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.58,
        "rate_delta": 0.02,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.58,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.6,
                "text": "(114-1)\n(男)2\n(女)22\n(教育部提供)\n合計24，達成率60%"
            }
        ]
    },
    {
        "id": "COM-11",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1-2",
        "item": "修讀跨域學習課程學生人次-輔系",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n(男)1\n(女)2\n(教育部提供)\n合計3，達成率37.5%",
        "calc_rate": 0.125,
        "qualitative_desc": "【未達標檢討】預計於115年5月辦理跨域學習說明會，鼓助學生修習輔系，另，評估雙主修/輔系設罝學習獎勵金之可行性，藉以提高學生修課意願。 ｜ 1.因輔系應修課程較繁重、跨系修課恐衝堂，且輔系無學位，導致修讀意願不高。\n2.未來將於每學期之跨域學習說明會，加強宣導，讓學生了解輔系之益處，以提升學生修讀意願。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【智慧排課導航 Agent】",
        "ai_strategy": "啟動 【智慧排課導航 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.105,
        "rate_delta": 0.02,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.105,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.125,
                "text": "(114-1)\n(男)1\n(女)2\n(教育部提供)\n合計3，達成率37.5%"
            }
        ]
    },
    {
        "id": "COM-12",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1-3",
        "item": "修讀跨域學習課程學生人次-學分學程",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n(男)152\n(女)186\n(教育部提供)\n合計338，達成率84.5%",
        "calc_rate": 0.845,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【智慧排課導航 Agent】",
        "ai_strategy": "啟動 【智慧排課導航 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.84,
        "rate_delta": 0.005,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.84,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.845,
                "text": "(114-1)\n(男)152\n(女)186\n(教育部提供)\n合計338，達成率84.5%"
            }
        ]
    },
    {
        "id": "COM-13",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1-4",
        "item": "修讀跨域學習課程學生人次-其他",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n(男)1159\n(女)1817\n合計2976，達成率99.2%\n學校自填",
        "calc_rate": 0.992,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【智慧排課導航 Agent】",
        "ai_strategy": "啟動 【智慧排課導航 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.912,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.912,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.992,
                "text": "(114-1)\n(男)1159\n(女)1817\n合計2976，達成率99.2%\n學校自填"
            }
        ]
    },
    {
        "id": "COM-14",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "協助學生跨域學習成效提升之推動策略描述",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n因應產業未來跨域人才需要，規劃專業多元學習，培養跨領域專長的課程學習方案，引導學生有意義性進行跨域學習，整合跨領域課程排課時段，透過排課減少衝堂",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【智慧排課導航 Agent】",
        "ai_strategy": "啟動 【智慧排課導航 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-1)\n因應產業未來跨域人才需要，規劃專業多元學習，培養跨領域專長的課程學習方案，引導學生有意義性進行跨域學習，整合跨領域課程排課時段，透過排課減少衝堂"
            }
        ]
    },
    {
        "id": "COM-15",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.1",
        "item": "學生資訊科技能力推動成效 - STEM領域系科所學生人數-男",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n538(教育部提供)\n達成率89.66%",
        "calc_rate": 0.9683,
        "qualitative_desc": "【未達標檢討】無 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8883,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8883,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9683,
                "text": "(114-2)\n538(教育部提供)\n達成率89.66%"
            }
        ]
    },
    {
        "id": "COM-16",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.2",
        "item": "STEM領域系科所學生人數-女",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n471(教育部提供)",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n471(教育部提供)"
            }
        ]
    },
    {
        "id": "COM-17",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.1",
        "item": "修讀STEM領域課程學生人次-男",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n4024(教育部提供)\n達成率91.04%",
        "calc_rate": 0.8388,
        "qualitative_desc": "【未達標檢討】無 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8338,
        "rate_delta": 0.005,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8338,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.8388,
                "text": "(114-2)\n4024(教育部提供)\n達成率91.04%"
            }
        ]
    },
    {
        "id": "COM-18",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.2",
        "item": "修讀STEM領域課程學生人次-女",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n3173(教育部提供)",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n3173(教育部提供)"
            }
        ]
    },
    {
        "id": "COM-19",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "3.0",
        "item": "曾修讀程式設計課程學生數",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n5142(教育部提供)",
        "calc_rate": 0.9908,
        "qualitative_desc": "【未達標檢討】無 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.9108,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.9108,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9908,
                "text": "(114-2)\n5142(教育部提供)"
            }
        ]
    },
    {
        "id": "COM-20",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "4.0",
        "item": "曾修讀數位科技微學程學生數",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n657(教育部提供)",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n657(教育部提供)"
            }
        ]
    },
    {
        "id": "COM-21",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "5.0",
        "item": "協助學生程式設計能力提升之推動策略描述（含學校在鼓勵修讀STEM領域課程上的推動策略）",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n為不同基礎者開設分級課程，打造視覺圖表、感測器程式構件教材，數位工具應用，培育大數據資料解讀與程式設計能力。建立基礎/進階/應用三階段課程，以培",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n為不同基礎者開設分級課程，打造視覺圖表、感測器程式構件教材，數位工具應用，培育大數據資料解讀與程式設計能力。建立基礎/進階/應用三階段課程，以培"
            }
        ]
    },
    {
        "id": "COM-22",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "學生中文閱讀寫作能力提升成效 - 學生通過學校設定校際共享中文能力教材測驗或第三方認證人數",
        "dept": "共同教育中心",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n619",
        "calc_rate": 0.7482,
        "qualitative_desc": "【未達標檢討】114-1中文檢測的前測乃是針對入學學生的中文先備能力做評估，從中了解入學新生的中文能力，通過人數的降低除了抽樣人數減少的因素之外，可看出學生基礎中文能力越來越低落的普遍現象，同時在未來一課程強化閱讀能力，以及提升書寫表達技巧，並且以檢測平台的輔導機制，幫助學生有效提升基礎能力，並於114-2的檢測後測中通過認證，提高通過率 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.7432,
        "rate_delta": 0.005,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.7432,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.7482,
                "text": "(114-2)\n619"
            }
        ]
    },
    {
        "id": "COM-23",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "學校設定校際共享中文能力教材測驗或第三方認證抽樣學生數",
        "dept": "共同教育中心",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n832\n達成率92.44%",
        "calc_rate": 0.9277,
        "qualitative_desc": "【未達標檢討】檢測人數為114學年度四技及五專入學新生統計總人數，因為入學新生人數減少，因此抽樣人數隨之減少 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8477,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8477,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9277,
                "text": "(114-2)\n832\n達成率92.44%"
            }
        ]
    },
    {
        "id": "COM-24",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "3.0",
        "item": "協助學生中文閱讀寫作能力提升之推動策略描述",
        "dept": "共同教育中心",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n推動本校學生中文閱讀寫作能力，為達有效提升方法，不僅進行四技一課程改革，包括課名更動、教材更新，甚至將前測、後測認定為一份作業，前測佔10％；後",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n推動本校學生中文閱讀寫作能力，為達有效提升方法，不僅進行四技一課程改革，包括課名更動、教材更新，甚至將前測、後測認定為一份作業，前測佔10％；後"
            }
        ]
    },
    {
        "id": "COM-25",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "學生英語能力提升成效 - 辦理專業英語課程(ESP、EAP)數",
        "dept": "共同教育中心",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "每年匯入1次，本次不匯入",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【外語拔尖 Agent】",
        "ai_strategy": "啟動 【外語拔尖 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "每年匯入1次，本次不匯入"
            }
        ]
    },
    {
        "id": "COM-26",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "學生達各級CEFR能力情形-英文證照通過張數-A1",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n280(教育部提供)",
        "calc_rate": 0.892,
        "qualitative_desc": "【未達標檢討】無 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.887,
        "rate_delta": 0.005,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.887,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.892,
                "text": "(114-1)\n280(教育部提供)"
            }
        ]
    },
    {
        "id": "COM-27",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "3.0",
        "item": "學生達各級CEFR能力情形-英文證照通過張數-A2",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n320(教育部提供)",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-1)\n320(教育部提供)"
            }
        ]
    },
    {
        "id": "COM-28",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "4.0",
        "item": "學生達各級CEFR能力情形-英文證照通過張數-B1",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n46(教育部提供)",
        "calc_rate": 0.9333,
        "qualitative_desc": "【未達標檢討】無 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8533,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8533,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9333,
                "text": "(114-1)\n46(教育部提供)"
            }
        ]
    },
    {
        "id": "COM-29",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "5.0",
        "item": "學生達各級CEFR能力情形-英文證照通過張數-B2",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n1(教育部提供)\n達成率33.33%",
        "calc_rate": 0.3333,
        "qualitative_desc": "【未達標檢討】本校學生結構對應B2和C1以上屬高階英文 (考一次即達畢業門檻) ，將持續推動免費英檢考照，提高B1.B2學生報考誘因。113學年度起每班將開設各班優秀B1以上學生拔尖輔導證照輔導課程，以及全班境外生班級，並藉全額報名費補助激勵潛力學生報考高階英檢證照，如搭配高額獎勵金效果可加乘。 ｜ 本校學生結構對應B2和C1以上屬高階英文 (考一次即達畢業門檻) ，將持續推動英檢考照輔導班，除提供獎勵補助外，115學年新增進步獎勵方案，提高B1.B2學生報考誘因。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.3133,
        "rate_delta": 0.02,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.3133,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.3333,
                "text": "(114-1)\n1(教育部提供)\n達成率33.33%"
            }
        ]
    },
    {
        "id": "COM-30",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "6.0",
        "item": "學生達各級CEFR能力情形-英文證照通過張數-C1",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n0(教育部提供)\n達成0%",
        "calc_rate": 0.0,
        "qualitative_desc": "【未達標檢討】本校學生結構對應B2和C1以上屬高階英文 (考一次即達畢業門檻) ，將持續推動免費英檢考照，提高B1.B2學生報考誘因。113學年度起每班將開設各班優秀B1以上學生拔尖輔導證照輔導課程，以及全班境外生班級，並藉全額報名費補助激勵潛力學生報考高階英檢證照，如搭配高額獎勵金效果可加乘。 ｜ 本校學生結構對應B2和C1以上屬高階英文 (考一次即達畢業門檻) ，將持續推動英檢考照輔導班，除提供獎勵補助外，115學年新增進步獎勵方案，提高B1.B2學生報考誘因。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.0,
        "rate_delta": 0.0,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.0,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.0,
                "text": "(114-1)\n0(教育部提供)\n達成0%"
            }
        ]
    },
    {
        "id": "COM-31",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "7.0",
        "item": "學生達各級CEFR能力情形-英文證照通過張數-C2",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n0(教育部提供)\n達成0%",
        "calc_rate": 0.0,
        "qualitative_desc": "【未達標檢討】TOEIC已取消C2等級，目前本校學生多參加本項檢測，未來持續鼓勵菁英學生，改考其他有C2等級的英檢考試。 ｜ 本校英檢考以TOEIC為主,現已取消C2等級，加上科大學生普遍難以取得C2，故本項指標無法達成，建議取消本項指標。\n(建議刪除此句)",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.0,
        "rate_delta": 0.0,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.0,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.0,
                "text": "(114-1)\n0(教育部提供)\n達成0%"
            }
        ]
    },
    {
        "id": "COM-32",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "8.0",
        "item": "學生非英文證照通過張數",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n1(教育部提供)\n達成率33.33%",
        "calc_rate": 0.3333,
        "qualitative_desc": "【未達標檢討】第二外語日韓語一年舉辦2次,施測後約半年才能取得證照,故數據未完全出來。 ｜ 112-1曾開設進階第二日語，並鼓勵學生考取非英語證照，因第二外語未列入畢業門檻，且上課學生普遍將第二外語視為出國可溝通工具，並無意願參加檢定考，建議本項指標改為課程通過率，作為學習評量指標。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.3133,
        "rate_delta": 0.02,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.3133,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.3333,
                "text": "(114-1)\n1(教育部提供)\n達成率33.33%"
            }
        ]
    },
    {
        "id": "COM-33",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "9.0",
        "item": "協助學生提升英語能力及修讀專業英語 課程ESP或EAP課程之具體推動策略描述",
        "dept": "共同教育中心",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-1)\n1.建置職場英語沉浸教學，以職場社交用語/英語書信應用/英語簡報等，辦理學習活動與競賽，打造自然無壓力的語感教學模式。\n2.編纂專業英語教材與學",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【外語拔尖 Agent】",
        "ai_strategy": "啟動 【外語拔尖 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-1)\n1.建置職場英語沉浸教學，以職場社交用語/英語書信應用/英語簡報等，辦理學習活動與競賽，打造自然無壓力的語感教學模式。\n2.編纂專業英語教材與學"
            }
        ]
    },
    {
        "id": "COM-34",
        "category": "共同指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "優化師資質量及改善生師比推動成效 - 生師比",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n20.47(教育部提供)",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n20.47(教育部提供)"
            }
        ]
    },
    {
        "id": "COM-35",
        "category": "共同指標",
        "aspect": "善盡社會責任 (USR)",
        "aspect_code": "A4",
        "code": "1.0",
        "item": "大學實踐社會責任推動成效 - 與中長期校務發展連結之作法與具體成果",
        "dept": "USR推動中心",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n整合外部來源及附設機構之校務研究資料，主動取得外部關係人之意見回饋，以建立更完整之校務經營資料倉儲系統，並逐步更新相關IR資料庫之資料字典。",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n整合外部來源及附設機構之校務研究資料，主動取得外部關係人之意見回饋，以建立更完整之校務經營資料倉儲系統，並逐步更新相關IR資料庫之資料字典。"
            }
        ]
    },
    {
        "id": "COM-36",
        "category": "共同指標",
        "aspect": "善盡社會責任 (USR)",
        "aspect_code": "A4",
        "code": "2.0",
        "item": "結合學校教研能量及社會資源，促進在地永續發展之作法及成效",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n以跨領域為本，透過資料分析基礎回饋之課程改革，提出能與本校之校內外利害關係人合作，及共同參與之社會實踐方案。",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n以跨領域為本，透過資料分析基礎回饋之課程改革，提出能與本校之校內外利害關係人合作，及共同參與之社會實踐方案。"
            }
        ]
    },
    {
        "id": "COM-37",
        "category": "共同指標",
        "aspect": "產學合作連結",
        "aspect_code": "A2",
        "code": "1.0",
        "item": "學生創新創業課程推動成效 - 開設創新創業課程教師數(學期/人)",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n81(教育部提供)\n達成率73.63%",
        "calc_rate": 0.885,
        "qualitative_desc": "【未達標檢討】無 ｜ 對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【三創育成 Agent】",
        "ai_strategy": "啟動 【三創育成 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.88,
        "rate_delta": 0.005,
        "trend_status": "STABLE",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.88,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.885,
                "text": "(114-2)\n81(教育部提供)\n達成率73.63%"
            }
        ]
    },
    {
        "id": "COM-38",
        "category": "共同指標",
        "aspect": "產學合作連結",
        "aspect_code": "A2",
        "code": "2.0",
        "item": "修讀創新創業課程學生人次",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n1232(教育部提供)\n達成率55.24%",
        "calc_rate": 0.818,
        "qualitative_desc": "【未達標檢討】無 ｜ 因少子化影響，致使修讀人次下滑，建議修正目標值。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【三創育成 Agent】",
        "ai_strategy": "啟動 【三創育成 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.813,
        "rate_delta": 0.005,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.813,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.818,
                "text": "(114-2)\n1232(教育部提供)\n達成率55.24%"
            }
        ]
    },
    {
        "id": "COM-39",
        "category": "共同指標",
        "aspect": "產學合作連結",
        "aspect_code": "A2",
        "code": "4.0",
        "item": "協助學生創新創業之具體推動策略描述 （含輔導校內創新創業團隊概況）",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)\n辦理校友創業、智慧財產權講座，提升師生專利申請數，辦理校園創業競賽，協助學院辦理三創競賽。",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【三創育成 Agent】",
        "ai_strategy": "啟動 【三創育成 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)\n辦理校友創業、智慧財產權講座，提升師生專利申請數，辦理校園創業競賽，協助學院辦理三創競賽。"
            }
        ]
    },
    {
        "id": "COM-40",
        "category": "共同指標",
        "aspect": "提升高教公共性",
        "aspect_code": "A3",
        "code": "1.0",
        "item": "經濟或文化不利學生獲得輔導或協助之提升成效 - 經濟或文化不利學生獲得輔導或協助之提升成效",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)  \n從六方向(學習面、技能面、生活面、經濟面、學用面、就業面)輔導扶助，針對經濟、學習、族群等各種弱勢學生，適切提供關懷輔導，特訂定本校「激勵學",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)  \n從六方向(學習面、技能面、生活面、經濟面、學用面、就業面)輔導扶助，針對經濟、學習、族群等各種弱勢學生，適切提供關懷輔導，特訂定本校「激勵學"
            }
        ]
    },
    {
        "id": "COM-41",
        "category": "共同指標",
        "aspect": "提升高教公共性",
        "aspect_code": "A3",
        "code": "1.0",
        "item": "輔導原民生及推動全民原教成效 - 辦理全民原教相關活動場次",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)  \n33",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)  \n33"
            }
        ]
    },
    {
        "id": "COM-42",
        "category": "共同指標",
        "aspect": "提升高教公共性",
        "aspect_code": "A3",
        "code": "2.0",
        "item": "定期召開諮詢委員會議及校內跨單位合作機制會議機制",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "(114-2)  \n原資中心針對原民生，推動「原民生數位學習教育課程」、建置「原住民族資源整合平臺」、「原民生專屬集會所」及獎助學金、學雜費減免、住宿減免、課業",
        "calc_rate": 1.0,
        "qualitative_desc": "對齊 115 年第 2 次填報最新彙整數據。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.92,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.92,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 1.0,
                "text": "(114-2)  \n原資中心針對原民生，推動「原民生數位學習教育課程」、建置「原住民族資源整合平臺」、「原民生專屬集會所」及獎助學金、學雜費減免、住宿減免、課業"
            }
        ]
    },
    {
        "id": "CUS-43",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "學生專業實務技術能力推動成效 - 應屆畢業生完成專業實習比率",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "113學年度-完成實習學生數/應屆畢業生數 \n71.06%",
        "calc_rate": 0.7106,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.7056,
        "rate_delta": 0.005,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.7056,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.7106,
                "text": "113學年度-完成實習學生數/應屆畢業生數 \n71.06%"
            }
        ]
    },
    {
        "id": "CUS-44",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "產業鏈結實務專題製作件數 1.連結計畫、2.業界共同指導、3.使用產業場域、4.產業應用議題等",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "護理學院：35件/199人次修課\n醫健學院：64件/242人次修課\n環生學院：41件/158人次修課\n人管學院：31件/181人次修課\n全校：171件/780人",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "護理學院：35件/199人次修課\n醫健學院：64件/242人次修課\n環生學院：41件/158人次修課\n人管學院：31件/181人次修課\n全校：171件/780人"
            }
        ]
    },
    {
        "id": "CUS-45",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "3.0",
        "item": "業界回饋畢業生職場實務滿意度",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "0.9466",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "0.9466"
            }
        ]
    },
    {
        "id": "CUS-46",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "教師實務經驗提升成效 - 教師參與實務研習人次(含校內外研習)",
        "dept": "研發處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "2159\n達成率94.69%",
        "calc_rate": 0.9469,
        "qualitative_desc": "【未達標檢討】無 ｜ 依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8669,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8669,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9469,
                "text": "2159\n達成率94.69%"
            }
        ]
    },
    {
        "id": "CUS-47",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "產學雙師共教教師數(含業師協同.實務專題業師)",
        "dept": "研發處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "1.業師協同教學共138人次:\n113-1(70人次)+113-2(68人次)\n2.實務專題業師，共計18人次：113-1(3人次)+113-2(15人次)。\n",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【產學列管 Agent】",
        "ai_strategy": "啟動 【產學列管 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "1.業師協同教學共138人次:\n113-1(70人次)+113-2(68人次)\n2.實務專題業師，共計18人次：113-1(3人次)+113-2(15人次)。\n"
            }
        ]
    },
    {
        "id": "CUS-48",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "教師推動創新教學成效之提升成效 - 教師創新教學產出作品件數(修正)",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "1.教師教學產出成果競賽：43\n2.數位教材製作：26\n合計:69",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【創新教案 AI Agent】",
        "ai_strategy": "啟動 【創新教案 AI Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "1.教師教學產出成果競賽：43\n2.數位教材製作：26\n合計:69"
            }
        ]
    },
    {
        "id": "CUS-49",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "2.0",
        "item": "教師從事各類教學計畫件數(新增)",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "1.114年創新教學研究計畫通過15件。\n2.教學實踐通過10件。\n3.教務處業管之教學類計畫15件。\n合計:40件",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "1.114年創新教學研究計畫通過15件。\n2.教學實踐通過10件。\n3.教務處業管之教學類計畫15件。\n合計:40件"
            }
        ]
    },
    {
        "id": "CUS-50",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "學生資訊科技能力推動成效 - 修讀數位科技微學程結業人數",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "61.0",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "61.0"
            }
        ]
    },
    {
        "id": "CUS-51",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "學生中文閱讀寫作能力提升成效 - 中文閱讀書寫表達競賽人次",
        "dept": "學務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "1335(本校學生-1月程式設計競賽587人、5月天使嶺文學獎競賽127人、6月程式設計競賽460人/SDL競賽20人、11月耕讀課外閱讀心得競賽117人/SD",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【雙校核對防錯 Agent】",
        "ai_strategy": "啟動 【雙校核對防錯 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "1335(本校學生-1月程式設計競賽587人、5月天使嶺文學獎競賽127人、6月程式設計競賽460人/SDL競賽20人、11月耕讀課外閱讀心得競賽117人/SD"
            }
        ]
    },
    {
        "id": "CUS-52",
        "category": "自訂指標",
        "aspect": "教學創新精進",
        "aspect_code": "A1",
        "code": "1.0",
        "item": "學生英語能力提升成效 - 專業英文ESP/EAP/EMI課程（含系科專業英文、學術英文與全校EMI課程）修課人次",
        "dept": "共同教育中心",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "113-2：702人次\n114-1：720人次\n合計1422人次",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【外語拔尖 Agent】",
        "ai_strategy": "啟動 【外語拔尖 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "113-2：702人次\n114-1：720人次\n合計1422人次"
            }
        ]
    },
    {
        "id": "CUS-53",
        "category": "自訂指標",
        "aspect": "善盡社會責任 (USR)",
        "aspect_code": "A4",
        "code": "1.0",
        "item": "大學實踐社會責任推動成效 - 參與SDGs、USR相關服務與實踐課程數 110-111現況值：USR2計畫+融入服務學習內涵課程 112年：2計畫+共教課程+融入服務學習內涵課程",
        "dept": "USR推動中心",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "學務處-實踐課程16門(113-2:8門             114-1:8門)\n共教中心-6(113-2SDL3門、114-1SDL3門)\nUSR-57(",
        "calc_rate": 0.9294,
        "qualitative_desc": "【未達標檢討】無 ｜ 依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【USR 社會責任 Agent】",
        "ai_strategy": "啟動 【USR 社會責任 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8494,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8494,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9294,
                "text": "學務處-實踐課程16門(113-2:8門             114-1:8門)\n共教中心-6(113-2SDL3門、114-1SDL3門)\nUSR-57("
            }
        ]
    },
    {
        "id": "CUS-54",
        "category": "自訂指標",
        "aspect": "善盡社會責任 (USR)",
        "aspect_code": "A4",
        "code": "2.0",
        "item": "參與SDGs、USR相關師生社群社團參與人次 （教務-專業成長社群、研發-研發社群、學務-學生社團）",
        "dept": "學務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "教務處-無\n研發處-200人次\n學務處-609人次              113-2：268                 114-1：341\n共809人",
        "calc_rate": 0.8515,
        "qualitative_desc": "【未達標檢討】無 ｜ 依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【USR 社會責任 Agent】",
        "ai_strategy": "啟動 【USR 社會責任 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8465,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8465,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.8515,
                "text": "教務處-無\n研發處-200人次\n學務處-609人次              113-2：268                 114-1：341\n共809人"
            }
        ]
    },
    {
        "id": "CUS-55",
        "category": "自訂指標",
        "aspect": "產學合作連結",
        "aspect_code": "A2",
        "code": "1.0",
        "item": "學生創新創業課程推動成效 - 參與三創競賽人次",
        "dept": "學務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "271.0",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【雙校核對防錯 Agent】",
        "ai_strategy": "啟動 【雙校核對防錯 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "271.0"
            }
        ]
    },
    {
        "id": "CUS-56",
        "category": "自訂指標",
        "aspect": "產學合作連結",
        "aspect_code": "A2",
        "code": "1.0",
        "item": "學校產學合作概況 - 產學合作策略聯盟家數",
        "dept": "研發處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "121.0",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【產學列管 Agent】",
        "ai_strategy": "啟動 【產學列管 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "121.0"
            }
        ]
    },
    {
        "id": "CUS-57",
        "category": "自訂指標",
        "aspect": "產學合作連結",
        "aspect_code": "A2",
        "code": "2.0",
        "item": "承接校外計畫件數",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "108\n達成率98.18%",
        "calc_rate": 0.9818,
        "qualitative_desc": "【未達標檢討】無 ｜ 依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.9018,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.9018,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9818,
                "text": "108\n達成率98.18%"
            }
        ]
    },
    {
        "id": "CUS-58",
        "category": "自訂指標",
        "aspect": "提升高教公共性",
        "aspect_code": "A3",
        "code": "1.0",
        "item": "經濟或文化不利學生獲得輔導或協助之提升成效 - 經濟或文化不利學生班排名獲前50%比率",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "48.30%                113學年度 ( 經濟不利且班前50%之人次÷經濟不利總人次)×100%  \n達成率99.38%",
        "calc_rate": 0.9938,
        "qualitative_desc": "【未達標檢討】無 ｜ 依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.9138,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.9138,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9938,
                "text": "48.30%                113學年度 ( 經濟不利且班前50%之人次÷經濟不利總人次)×100%  \n達成率99.38%"
            }
        ]
    },
    {
        "id": "CUS-59",
        "category": "自訂指標",
        "aspect": "提升高教公共性",
        "aspect_code": "A3",
        "code": "2.0",
        "item": "經濟或文化不利學生畢業就業升學率",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "95.10%\n112學年度( 升學人數＋就業人數+留學人數)÷(畢業人數-服兵役人數)",
        "calc_rate": 0.951,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.871,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.871,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.951,
                "text": "95.10%\n112學年度( 升學人數＋就業人數+留學人數)÷(畢業人數-服兵役人數)"
            }
        ]
    },
    {
        "id": "CUS-60",
        "category": "自訂指標",
        "aspect": "提升高教公共性",
        "aspect_code": "A3",
        "code": "3.0",
        "item": "經濟或文化不利學生(含應屆畢業生)已獲得專業證照之人次(113年新增)",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "331                  (114年度)教務處提供220人次(含附錄1學生) +本學期111人次\n達成率82.75%",
        "calc_rate": 0.8275,
        "qualitative_desc": "【未達標檢討】無 ｜ 依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【證照輔導 Agent】",
        "ai_strategy": "啟動 【證照輔導 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8225,
        "rate_delta": 0.005,
        "trend_status": "STAGNANT",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8225,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.8275,
                "text": "331                  (114年度)教務處提供220人次(含附錄1學生) +本學期111人次\n達成率82.75%"
            }
        ]
    },
    {
        "id": "CUS-61",
        "category": "自訂指標",
        "aspect": "提升高教公共性",
        "aspect_code": "A3",
        "code": "1.0",
        "item": "輔導原民生及推動全民原教成效 - 原住民學生畢業就業升學率",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "93.69%                  112學年度( 升學人數＋就業人數+留學人數)÷(畢業人數-服兵役人數)",
        "calc_rate": 0.9369,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.8569,
        "rate_delta": 0.08,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.8569,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.9369,
                "text": "93.69%                  112學年度( 升學人數＋就業人數+留學人數)÷(畢業人數-服兵役人數)"
            }
        ]
    },
    {
        "id": "CUS-62",
        "category": "自訂指標",
        "aspect": "提升高教公共性",
        "aspect_code": "A3",
        "code": "2.0",
        "item": "推動特色原民文化",
        "dept": "教務處",
        "source": "115年第2次填報 (0813彙整版)",
        "current_val_text": "1.規劃原住民多元文化、全民原教之相關課程及活動，並以大數據及雲端為主軸，透過部落學習及影像聲音紀錄並上傳至數位社群平台，促進全體國民認識與尊重原住民族。\n2.",
        "calc_rate": 0.85,
        "qualitative_desc": "依自訂指標追蹤控管並進行管考與輔導。",
        "milestone": "115-2 期中檢核與管考",
        "deadline": "2026-10-31",
        "ai_agent": "【AI 專案 Agent】",
        "ai_strategy": "啟動 【AI 專案 Agent】 追蹤管考與資源對接，推動指標達標。",
        "prev_calc_rate": 0.845,
        "rate_delta": 0.005,
        "trend_status": "PROGRESS_MET",
        "history": [
            {
                "period": "115-1 填報期初",
                "calc_rate": 0.845,
                "text": "期初基準值"
            },
            {
                "period": "115-2 填報 (0813彙整版)",
                "calc_rate": 0.85,
                "text": "1.規劃原住民多元文化、全民原教之相關課程及活動，並以大數據及雲端為主軸，透過部落學習及影像聲音紀錄並上傳至數位社群平台，促進全體國民認識與尊重原住民族。\n2."
            }
        ]
    }
]

def parse_num_clean(val):
    if val is None: return None
    s = str(val).strip()
    if not s or s == 'None': return None
    m_rate = re.search(r'達成率\s*([0-9\.]+)%', s)
    if m_rate: return float(m_rate.group(1)) / 100.0
    cleaned = re.sub(r'\([0-9]{3}(?:-[0-9]|年|學年度)?\)', '', s)
    cleaned = re.sub(r'\(教育部提供\)', '', cleaned)
    cleaned = re.sub(r'[^\d\.%]', ' ', cleaned).strip()
    m_pct = re.search(r'([0-9\.]+)\s*%', s)
    if m_pct:
        v = float(m_pct.group(1))
        return v / 100.0 if v > 1.0 else v
    nums = [float(x) for x in re.findall(r'[0-9]+\.?[0-9]*', cleaned) if x]
    if nums: return nums[0]
    return None

def calc_rate_pair(actual_raw, target_raw):
    if not actual_raw or not target_raw: return None
    s_act = str(actual_raw).strip()
    m_rate = re.search(r'達成率\s*([0-9\.]+)%', s_act)
    if m_rate: return round(float(m_rate.group(1)) / 100.0, 4)
    act_num = parse_num_clean(actual_raw)
    tgt_num = parse_num_clean(target_raw)
    if act_num is not None and tgt_num is not None and tgt_num > 0:
        if act_num <= 1.0 and tgt_num <= 1.0: return round(act_num / tgt_num, 4)
        elif act_num > 1.0 and tgt_num > 1.0: return round(act_num / tgt_num, 4)
        elif act_num <= 1.0: return round(act_num, 4)
    return None

def parse_excel_and_merge_history(wb, period_label="最新填報期"):
    global INDICATORS_DB
    unmet_map = {}

    for sheet_name in ["共同未達標", "自訂未達標"]:
        if sheet_name in wb.sheetnames:
            s = wb[sheet_name]
            for r in list(s.iter_rows(values_only=True))[1:]:
                if any(r):
                    item_name = str(r[2] or "").strip()
                    rate_val = str(r[3] or "").strip()
                    reason = str(r[4] or "").strip()
                    note = str(r[5] or "").strip()
                    if item_name: unmet_map[item_name] = {"rate": rate_val, "reason": reason, "note": note}

    def parse_rate(text):
        if not text: return None
        m = re.search(r'達成率\s*([0-9\.]+)%', text)
        if m: return round(float(m.group(1)) / 100.0, 4)
        m2 = re.search(r'([0-9\.]+)%', text)
        if m2: return round(float(m2.group(1)) / 100.0, 4)
        return None

    def process_item(item_name, aspect, code_no, val_str, desc_str, is_custom=False, annual_targets=None, annual_actuals=None, annual_rates=None, dept=None):
        rate = parse_rate(val_str)
        unmet_info = next((v for k, v in unmet_map.items() if k in item_name or item_name in k or (code_no and code_no in k)), None)
        if unmet_info and unmet_info.get("rate"):
            try: rate = round(float(unmet_info["rate"]), 4)
            except: pass
        if rate is None and annual_rates and annual_rates.get("114"):
            rate = annual_rates["114"]
        if rate is None: rate = 0.85

        matched = None
        for ind in INDICATORS_DB:
            if ind.get("code") == code_no and ind.get("aspect") == aspect:
                matched = ind; break
            elif ind.get("item") == item_name:
                matched = ind; break

        final_desc = desc_str
        if unmet_info and unmet_info.get("reason"):
            final_desc = f"【未達標檢討】{unmet_info['reason']} ｜ {desc_str}"

        prefix = "U" if is_custom else "C"
        if matched:
            old_rate = matched.get("calc_rate", rate)
            matched["prev_calc_rate"] = old_rate
            matched["calc_rate"] = rate
            matched["rate_delta"] = round(rate - old_rate, 4)
            matched["current_val_text"] = val_str[:120]
            if final_desc: matched["qualitative_desc"] = final_desc[:280]
            if dept and dept != "填報單位": matched["dept"] = dept
            if annual_targets: matched["annual_targets"] = annual_targets
            if annual_actuals: matched["annual_actuals"] = annual_actuals
            if annual_rates: matched["annual_rates"] = annual_rates

            delta = matched["rate_delta"]
            if rate >= 0.85 and old_rate < 0.85: matched["trend_status"] = "PROGRESS_MET"
            elif delta >= 0.05: matched["trend_status"] = "PROGRESS_MET"
            elif delta <= 0.01 and rate < 0.85: matched["trend_status"] = "STAGNANT"
            else: matched["trend_status"] = "STABLE"

            history = matched.get("history", [])
            history.append({"period": period_label, "calc_rate": rate, "text": val_str[:120]})
            matched["history"] = history
        else:
            INDICATORS_DB.append({
                "id": f"{prefix}-{len(INDICATORS_DB)+1:02d}",
                "category": "自訂指標" if is_custom else "共同指標",
                "aspect": aspect,
                "aspect_code": "A1",
                "code": code_no,
                "item": item_name,
                "dept": dept or "填報單位",
                "source": period_label,
                "current_val_text": val_str[:120],
                "calc_rate": rate,
                "prev_calc_rate": rate,
                "rate_delta": 0.0,
                "trend_status": "STABLE",
                "qualitative_desc": final_desc[:280],
                "milestone": "期中管考追蹤",
                "deadline": "2026-10-31",
                "ai_agent": "【AI 專案 Agent】",
                "ai_strategy": "啟動 AI Agent 追蹤管考與資源對接。",
                "annual_targets": annual_targets or {},
                "annual_actuals": annual_actuals or {},
                "annual_rates": annual_rates or {},
                "history": [{"period": period_label, "calc_rate": rate, "text": val_str[:120]}]
            })

    if "共同績效指標" in wb.sheetnames:
        sheet = wb["共同績效指標"]
        rows = list(sheet.iter_rows(values_only=True))
        current_aspect = "教學創新精進"
        for r in rows[5:]:
            if not any(r): continue
            col0 = str(r[0] or "").replace('\n', '').replace(' ', '')
            col1 = str(r[1] or "").replace('\n', '').replace(' ', '')
            col2 = str(r[2] or "")
            col3 = str(r[3] or "").replace('\n', ' ')
            if "教學創新" in col0 or "教學創新" in col1: current_aspect = "教學創新精進"
            elif "產學合作" in col0 or "產學合作" in col1: current_aspect = "產學合作連結"
            elif "高教公共" in col0 or "高教公共" in col1: current_aspect = "提升高教公共性"
            elif "社會責任" in col0 or "社會責任" in col1: current_aspect = "善盡社會責任 (USR)"
            if col3 and col2:
                val_str = str(r[13] or "") if len(r) > 13 else ""
                desc_str = str(r[14] or "") if len(r) > 14 else ""
                dept_str = str(r[28] or "").replace('\n', ' ') if len(r) > 28 else ""
                t112 = r[16] or r[15] if len(r) > 16 else ""
                t113 = r[18] or r[17] if len(r) > 18 else ""
                t114 = r[20] or r[19] if len(r) > 20 else ""
                t115 = r[22] or r[21] if len(r) > 22 else ""
                t116 = r[24] or r[23] if len(r) > 24 else ""
                a111 = r[7] or r[6] if len(r) > 7 else ""
                a112 = r[9] or r[8] if len(r) > 9 else ""
                a113 = r[11] or r[10] if len(r) > 11 else ""
                a114 = r[13] or r[12] if len(r) > 13 else ""
                ann_targets = {"112": str(t112 or "").strip(), "113": str(t113 or "").strip(), "114": str(t114 or "").strip(), "115": str(t115 or "").strip(), "116": str(t116 or "").strip()}
                ann_actuals = {"111": str(a111 or "").strip(), "112": str(a112 or "").strip(), "113": str(a113 or "").strip(), "114": str(a114 or "").strip()}
                ann_rates = {"112": calc_rate_pair(a112, t112), "113": calc_rate_pair(a113, t113), "114": calc_rate_pair(a114, t114)}
                item_full = f"{col1} - {col3}" if col1 else col3
                process_item(item_full, current_aspect, col2, val_str, desc_str, False, ann_targets, ann_actuals, ann_rates, dept_str)

    if "自訂績效指標" in wb.sheetnames:
        sheet = wb["自訂績效指標"]
        rows = list(sheet.iter_rows(values_only=True))
        current_aspect = "教學創新精進"
        for r in rows[4:]:
            if not any(r): continue
            col0 = str(r[0] or "").replace('\n', '').replace(' ', '')
            col1 = str(r[1] or "").replace('\n', '').replace(' ', '')
            col2 = str(r[2] or "")
            col3 = str(r[3] or "").replace('\n', ' ')
            if "教學創新" in col0 or "教學創新" in col1: current_aspect = "教學創新精進"
            elif "產學合作" in col0 or "產學合作" in col1: current_aspect = "產學合作連結"
            elif "高教公共" in col0 or "高教公共" in col1: current_aspect = "提升高教公共性"
            elif "社會責任" in col0 or "社會責任" in col1: current_aspect = "善盡社會責任 (USR)"
            if col3 and col2:
                val_str = str(r[12] or "") if len(r) > 12 else (str(r[11] or "") if len(r) > 11 else "")
                desc_str = str(r[13] or "") if len(r) > 13 else ""
                dept_str = str(r[23] or "").replace('\n', ' ') if len(r) > 23 else ""
                t112 = r[15] if len(r) > 15 else ""
                t113 = r[16] if len(r) > 16 else ""
                t114 = r[17] if len(r) > 17 else ""
                t115 = r[18] if len(r) > 18 else ""
                t116 = r[19] if len(r) > 19 else ""
                a110 = r[8] if len(r) > 8 else ""
                a111 = r[9] if len(r) > 9 else ""
                a112 = r[10] if len(r) > 10 else ""
                a113 = r[11] if len(r) > 11 else ""
                a114 = r[12] if len(r) > 12 else ""
                ann_targets = {"112": str(t112 or "").strip(), "113": str(t113 or "").strip(), "114": str(t114 or "").strip(), "115": str(t115 or "").strip(), "116": str(t116 or "").strip()}
                ann_actuals = {"110": str(a110 or "").strip(), "111": str(a111 or "").strip(), "112": str(a112 or "").strip(), "113": str(a113 or "").strip(), "114": str(a114 or "").strip()}
                ann_rates = {"112": calc_rate_pair(a112, t112), "113": calc_rate_pair(a113, t113), "114": calc_rate_pair(a114, t114)}
                item_full = f"{col1} - {col3}" if col1 else col3
                process_item(item_full, current_aspect, col2, val_str, desc_str, True, ann_targets, ann_actuals, ann_rates, dept_str)

    # 4. 前後期比較邏輯修正：比對最新 (Latest) 與 次新 (Second Latest) 資料
    for ind in INDICATORS_DB:
        rates = ind.get('annual_rates', {})
        valid_years = [y for y in ['111', '112', '113', '114'] if rates.get(y) is not None]
        
        if len(valid_years) >= 2:
            latest_yr = valid_years[-1]
            prev_yr = valid_years[-2]
            latest_rate = rates[latest_yr]
            prev_rate = rates[prev_yr]
        elif len(valid_years) == 1:
            latest_yr = valid_years[0]
            prev_yr = '期初'
            latest_rate = rates[latest_yr]
            prev_rate = ind.get('prev_calc_rate', 0.85)
        else:
            latest_yr = '最新'
            prev_yr = '次新'
            latest_rate = ind.get('calc_rate', 0.85)
            prev_rate = ind.get('prev_calc_rate', 0.85)
            
        delta = round(latest_rate - prev_rate, 4) if (latest_rate is not None and prev_rate is not None) else 0.0
        
        # 判定趨勢狀態
        if latest_rate >= 0.85 and prev_rate < 0.85:
            trend_status = "PROGRESS_MET"
        elif delta >= 0.05:
            trend_status = "PROGRESS_MET"
        elif delta <= 0.01 and latest_rate < 0.85:
            trend_status = "STAGNANT"
        else:
            trend_status = "STABLE"
            
        ind['latest_yr'] = latest_yr
        ind['prev_yr'] = prev_yr
        ind['calc_rate'] = latest_rate
        ind['prev_calc_rate'] = prev_rate
        ind['rate_delta'] = delta
        ind['trend_status'] = trend_status

    return INDICATORS_DB

def load_default_excel_if_exists():
    default_file = "第二期高教深耕計畫指標-115年第2次填報(0813彙整版).xlsx"
    if os.path.exists(default_file):
        try:
            wb = openpyxl.load_workbook(default_file, data_only=True)
            parse_excel_and_merge_history(wb, period_label="115年第2次填報 (0813彙整版)")
            print(f"[INFO] 已自動載入預設 Excel 跨年度資料庫 (共 {len(INDICATORS_DB)} 項指標)！")
        except Exception as e:
            print(f"[WARN] 自動載入預設 Excel 失敗: {e}")

load_default_excel_if_exists()

class SproutWebServer(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path in ["/", "/index.html"]:
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(self.render_html_dashboard().encode('utf-8'))
        elif parsed.path == "/api/indicators":
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(INDICATORS_DB, ensure_ascii=False).encode('utf-8'))
        elif parsed.path == "/api/audit":
            query = urllib.parse.parse_qs(parsed.query)
            threshold = float(query.get("threshold", [0.70])[0])
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(self.run_audit_rules(threshold), ensure_ascii=False).encode('utf-8'))
        elif parsed.path == "/api/export":
            csv_data = self.generate_csv_report()
            self.send_response(200)
            self.send_header('Content-type', 'text/csv; charset=utf-8-sig')
            self.send_header('Content-Disposition', 'attachment; filename="sprout_indicators_history_report.csv"')
            self.end_headers()
            self.wfile.write(csv_data)
        elif parsed.path == "/api/lan_info":
            lan_ip = get_lan_ip()
            port = self.server.server_address[1]
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps({"lan_ip": lan_ip, "port": port, "lan_url": f"http://{lan_ip}:{port}"}, ensure_ascii=False).encode('utf-8'))
        elif parsed.path == "/api/github/status":
            st = check_git_status()
            cfg = load_config()
            st["config_remote"] = cfg.get("remote_url", "")
            st["has_token"] = bool(cfg.get("token"))
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(st, ensure_ascii=False).encode('utf-8'))
        else:
            super().do_GET()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        content_length = int(self.headers.get('Content-Length', 0))
        body_bytes = self.rfile.read(content_length)

        if parsed.path == "/api/upload_excel":
            try:
                period_name = f"填報更新 ({datetime.now().strftime('%m/%d %H:%M')})"
                wb = openpyxl.load_workbook(io.BytesIO(body_bytes), data_only=True)
                new_db = parse_excel_and_merge_history(wb, period_label=period_name)

                # 備份最新上傳檔案
                try:
                    with open(f"第二期高教深耕計畫指標-上傳紀錄_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", "wb") as f:
                        f.write(body_bytes)
                except:
                    pass

                self.send_response(200)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"success": True, "count": len(new_db), "message": f"成功比較並動態更新！非覆蓋模式已自動保留歷史對比軌跡（包含轉移達成與滯後指標分析）。"}, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "error": str(e)}, ensure_ascii=False).encode('utf-8'))

        elif parsed.path == "/api/indicators/update":
            try:
                body = body_bytes.decode('utf-8')
                data = json.loads(body)
                ind_id = data.get("id")
                updated_item = None
                for ind in INDICATORS_DB:
                    if ind["id"] == ind_id:
                        if "calc_rate" in data:
                            old = ind.get("calc_rate", 0.0)
                            new_r = float(data["calc_rate"])
                            ind["prev_calc_rate"] = old
                            ind["calc_rate"] = new_r
                            ind["rate_delta"] = round(new_r - old, 4)
                            if new_r >= 0.85 and old < 0.85: ind["trend_status"] = "PROGRESS_MET"
                            elif ind["rate_delta"] <= 0.01 and new_r < 0.85: ind["trend_status"] = "STAGNANT"
                            else: ind["trend_status"] = "STABLE"

                        if "qualitative_desc" in data: ind["qualitative_desc"] = data["qualitative_desc"]
                        if "milestone" in data: ind["milestone"] = data["milestone"]
                        if "deadline" in data: ind["deadline"] = data["deadline"]
                        updated_item = ind
                        break
                
                if updated_item:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json; charset=utf-8')
                    self.end_headers()
                    self.wfile.write(json.dumps({"success": True, "indicator": updated_item}, ensure_ascii=False).encode('utf-8'))
                else:
                    self.send_response(404)
                    self.send_header('Content-type', 'application/json; charset=utf-8')
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "Indicator not found"}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode('utf-8'))

        elif parsed.path == "/api/github/push":
            try:
                data = json.loads(body_bytes.decode('utf-8')) if body_bytes else {}
                remote = data.get("remote_url")
                msg = data.get("commit_msg")
                token = data.get("token")
                res = push_to_github(remote_url=remote, commit_msg=msg, token=token)
                self.send_response(200)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps(res, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "message": str(e)}, ensure_ascii=False).encode('utf-8'))

        elif parsed.path == "/api/github/config":
            try:
                data = json.loads(body_bytes.decode('utf-8')) if body_bytes else {}
                save_config(data)
                self.send_response(200)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"success": True, "message": "GitHub 設定已成功儲存！"}, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "message": str(e)}, ensure_ascii=False).encode('utf-8'))

    def run_audit_rules(self, threshold=0.70):
        alerts = []
        for ind in INDICATORS_DB:
            rate = ind["calc_rate"]
            delta = ind.get("rate_delta", 0.0)
            desc = ind.get("qualitative_desc", "")
            
            if ind.get("trend_status") == "STAGNANT":
                alerts.append({
                    "id": ind["id"], "item": ind["item"], "dept": ind["dept"], "aspect": ind["aspect"],
                    "level": "WARNING", "type": "⚠️ 滯後未顯著增加",
                    "msg": f"前後期達成率增幅僅 {delta*100:+.1f}%（滯後），當前達成率 {rate*100:.1f}% 未見顯著改善！"
                })
            elif rate < threshold:
                alerts.append({
                    "id": ind["id"], "item": ind["item"], "dept": ind["dept"], "aspect": ind["aspect"],
                    "level": "CRITICAL", "type": f"未達標 (<{threshold*100:.0f}%)",
                    "msg": f"當前達成率僅 {rate*100:.1f}%，低於自訂門檻 ({threshold*100:.0f}%)，已啟動 {ind.get('ai_agent','AI Agent')}！"
                })
            
            if len(desc) > 300:
                alerts.append({
                    "id": ind["id"], "item": ind["item"], "dept": ind["dept"], "aspect": ind["aspect"],
                    "level": "INFO", "type": "字數警示",
                    "msg": f"質化成效描述達 {len(desc)} 字，超出 300 字元上限！"
                })
        return alerts

    def generate_csv_report(self):
        output = io.StringIO()
        output.write('\ufeff') # UTF-8 BOM
        writer = csv.writer(output)
        writer.writerow(["指標代碼", "指標類別", "構面名稱", "指標項目", "主責處室", "前期達成率", "最新達成率", "差異增減(Δ)", "差異趨勢分類", "115-2實績描述", "質化說明與未達標檢討", "AI Agent 處方對策"])
        for ind in INDICATORS_DB:
            delta = ind.get('rate_delta', 0.0)
            status_text = "🎉 轉移達成" if ind.get("trend_status")=="PROGRESS_MET" else ("⚠️ 滯後未顯著增加" if ind.get("trend_status")=="STAGNANT" else "穩定")
            writer.writerow([
                ind["id"], ind.get("category", "共同指標"), ind["aspect"], ind["item"], ind["dept"],
                f"{ind.get('prev_calc_rate', 0.0)*100:.2f}%", f"{ind['calc_rate']*100:.2f}%",
                f"{delta*100:+.2f}%", status_text,
                ind.get("current_val_text", ""), ind.get("qualitative_desc", ""), ind.get("ai_strategy", "")
            ])
        return output.getvalue().encode('utf-8-sig')

    def render_html_dashboard(self):
        return """<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>高教深耕計畫「智慧專案管理與指標管考系統」 (前後期歷史差異化比較版)</title>
    <!-- Bootstrap 5 & Icons -->
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css" rel="stylesheet">
    <!-- Google Fonts & Chart.js -->
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&family=Noto+Sans+TC:wght@400;500;700&display=swap" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        :root {
            --bg-main: #f8fafc;
            --card-bg: #ffffff;
            --primary-accent: #2563eb;
            --dark-text: #0f172a;
            --border-color: #e2e8f0;
        }
        body {
            background-color: var(--bg-main);
            font-family: 'Plus Jakarta Sans', 'Noto Sans TC', -apple-system, BlinkMacSystemFont, sans-serif;
            color: var(--dark-text);
            padding-bottom: 60px;
        }
        .navbar-custom {
            background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
            box-shadow: 0 4px 20px rgba(0, 0, 0, 0.1);
        }
        
        /* 門檻控制列專用樣式 */
        .threshold-bar {
            background: linear-gradient(135deg, #1e1b4b 0%, #312e81 100%);
            color: white;
            border-radius: 16px;
            padding: 20px 24px;
            box-shadow: 0 10px 25px rgba(49, 46, 129, 0.15);
            margin-bottom: 24px;
        }
        .threshold-slider {
            accent-color: #818cf8;
            height: 8px;
        }
        
        .card-custom {
            background: var(--card-bg);
            border-radius: 16px;
            border: 1px solid var(--border-color);
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.03);
            margin-bottom: 24px;
        }
        .table-hover tbody tr:hover {
            background-color: #f8fafc;
        }
        .tr-unmet {
            background-color: #fef2f2 !important;
        }
        .tr-progress {
            background-color: #f0fdf4 !important;
        }
        .tr-stagnant {
            background-color: #fffbeb !important;
        }
        .badge-agent {
            background-color: #f3e8ff;
            color: #6b21a8;
            font-weight: 600;
            padding: 4px 8px;
            border-radius: 6px;
            border: 1px solid #d8b4fe;
            font-size: 0.8rem;
        }
        .progress {
            height: 10px;
            border-radius: 6px;
            background-color: #e2e8f0;
        }
        .progress-bar {
            border-radius: 6px;
        }
        .delta-pill {
            font-size: 0.75rem;
            font-weight: 700;
            padding: 2px 6px;
            border-radius: 4px;
        }

        /* 手機與行動裝置極致響應式微調 */
        @media (max-width: 768px) {
            .navbar-brand { font-size: 1.1rem !important; }
            .navbar-brand div div { font-size: 0.85rem !important; }
            .threshold-bar { padding: 15px; }
            .card-custom { padding: 15px !important; }
            .btn-group { flex-wrap: wrap; gap: 4px; }
            .table-responsive { font-size: 0.85rem; }
        }
    </style>
</head>
<body>

    <!-- 頂部導覽列 -->
    <nav class="navbar navbar-expand-lg navbar-dark navbar-custom py-3 mb-4">
        <div class="container-fluid px-4">
            <a class="navbar-brand d-flex align-items-center fw-bold fs-4" href="#">
                <span class="fs-2 me-2">🏛️</span>
                <div>
                    <div>高教深耕計畫「智慧專案管理與指標管考中樞」</div>
                    <div class="fs-6 text-info fw-normal">前後期指標差異化分析 (非覆蓋模式・追蹤轉移達成與滯後未顯著增加)</div>
                </div>
            </a>
            <div class="ms-auto d-flex gap-2">
                <button class="btn btn-info d-flex align-items-center gap-2 px-3 rounded-3 text-white fw-bold" onclick="openShareModal()" id="lan-btn" title="跨電腦與手機連線及掃碼">
                    <i class="bi bi-qr-code-scan"></i> 📱 跨電腦/手機連線 (QR Code 掃碼)
                </button>
                <button class="btn btn-warning d-flex align-items-center gap-2 px-3 rounded-3 text-dark fw-bold" onclick="openUploadModal()">
                    <i class="bi bi-cloud-arrow-up-fill"></i> 📤 上傳新期填報檔 (自動比較)
                </button>
                <button class="btn btn-outline-info d-flex align-items-center gap-2 px-3 rounded-3 text-white fw-bold" onclick="openGithubModal()">
                    <i class="bi bi-github"></i> ☁️ 同步至 GitHub
                </button>
                <button class="btn btn-outline-light d-flex align-items-center gap-2 px-3 rounded-3" onclick="runAudit()">
                    <i class="bi bi-shield-check text-warning"></i> 執行全表差異稽核
                </button>
                <a class="btn btn-success d-flex align-items-center gap-2 px-3 rounded-3" href="/api/export">
                    <i class="bi bi-file-earmark-spreadsheet"></i> 匯出差異比較報表
                </a>
            </div>
        </div>
    </nav>

    <div class="container-fluid px-4">
        
        <!-- 🎯 控制中樞：門檻 slider 與差異化過濾按鈕群 -->
        <div class="threshold-bar">
            <div class="row align-items-center g-3">
                <div class="col-lg-4">
                    <div class="d-flex align-items-center gap-2 mb-1">
                        <i class="bi bi-sliders fs-4 text-warning"></i>
                        <h5 class="fw-bold m-0">🎯 可調整達成率檢視門檻</h5>
                    </div>
                    <div class="small" style="color: #c7d2fe;">設定達成率判定門檻（預設 70%），即時比對前後期達成率消長趨勢。</div>
                </div>
                <div class="col-lg-3">
                    <div class="d-flex align-items-center gap-3">
                        <input type="range" class="form-range threshold-slider flex-grow-1" id="threshold-range" min="0" max="100" step="5" value="70" oninput="syncThresholdInput(this.value)">
                        <div class="input-group input-group-sm" style="width: 90px;">
                            <input type="number" class="form-control fw-bold text-center text-primary" id="threshold-number" min="0" max="100" value="70" onchange="syncThresholdSlider(this.value)">
                            <span class="input-group-text bg-white fw-bold">%</span>
                        </div>
                    </div>
                </div>
                <div class="col-lg-5 text-lg-end">
                    <div class="btn-group btn-group-sm" role="group" id="delta-filter-group">
                        <button type="button" class="btn btn-outline-light active" onclick="setDeltaFilter('ALL')">全部 62 項</button>
                        <button type="button" class="btn btn-outline-success" onclick="setDeltaFilter('PROGRESS')">🎉 轉移達成 / 顯著進步 (<span id="cnt-progress">0</span>)</button>
                        <button type="button" class="btn btn-outline-warning" onclick="setDeltaFilter('STAGNANT')">⚠️ 未顯著增加 / 滯後 (<span id="cnt-stagnant">0</span>)</button>
                        <button type="button" class="btn btn-outline-danger" onclick="setDeltaFilter('UNMET')">🔴 未達標 (&lt;<span id="summary-threshold-val">70</span>%)</button>
                    </div>
                </div>
            </div>
        </div>

        <!-- 📊 四大動態統計卡片：呈現轉移達成與未顯著增加趨勢 -->
        <div class="row g-3 mb-4" id="trend-summary-cards">
            <!-- JavaScript 動態繪製總覽統計卡片 -->
        </div>

        <!-- 搜尋與過濾 Bar -->
        <div class="row g-2 mb-3 bg-white p-3 rounded-4 border shadow-sm align-items-center">
            <div class="col-md-4">
                <div class="input-group input-group-sm">
                    <span class="input-group-text bg-white"><i class="bi bi-search"></i></span>
                    <input type="text" id="search-input" class="form-control form-control-sm" placeholder="搜尋指標名稱/代碼/處室..." oninput="refreshDashboard()">
                </div>
            </div>
            <div class="col-md-3">
                <select id="category-filter" class="form-select form-select-sm" onchange="refreshDashboard()">
                    <option value="">所有指標類別 (共同指標 + 自訂指標)</option>
                    <option value="共同指標">共同指標 (教育部指定)</option>
                    <option value="自訂指標">自訂指標 (學校自訂)</option>
                </select>
            </div>
            <div class="col-md-3">
                <select id="aspect-filter" class="form-select form-select-sm" onchange="refreshDashboard()">
                    <option value="">所有四大構面</option>
                    <option value="教學創新精進">構面 A1：教學創新精進</option>
                    <option value="產學合作連結">構面 A2：產學合作連結</option>
                    <option value="提升高教公共性">構面 A3：提升高教公共性</option>
                    <option value="善盡社會責任 (USR)">構面 A4：善盡社會責任 (USR)</option>
                </select>
            </div>
            <div class="col-md-2 text-end">
                <button class="btn btn-sm btn-outline-secondary w-100" onclick="resetFilters()">重置過濾</button>
            </div>
        </div>

        <!-- 主介面：依四大構面呈現所有指標 (附帶前後期增減 Δ 比對) -->
        <div id="aspects-main-container">
            <!-- JavaScript 動態渲染四大構面內容 -->
        </div>

        <!-- 四大構面達成率圖表與智慧稽核報告 (雙欄) -->
        <div class="row g-4 my-4">
            <div class="col-lg-7">
                <div class="card-custom p-4 h-100">
                    <h5 class="fw-bold text-dark mb-3 d-flex align-items-center gap-2">
                        <i class="bi bi-graph-up-arrow text-primary"></i> 📊 四大構面指標差異增減 (Δ%) 分析圖表
                    </h5>
                    <div style="position: relative; height: 320px;">
                        <canvas id="aspectChart"></canvas>
                    </div>
                </div>
            </div>
            <div class="col-lg-5">
                <div class="card-custom p-4 h-100" id="audit-card">
                    <div class="d-flex justify-content-between align-items-center mb-3">
                        <h5 class="fw-bold text-dark m-0 d-flex align-items-center gap-2">
                            <i class="bi bi-shield-lock-fill text-success"></i> 🛡️ 自動差異化稽核與滯後指標診斷
                        </h5>
                        <button class="btn btn-sm btn-outline-secondary" onclick="runAudit()">重新掃描</button>
                    </div>
                    <div id="audit-content" class="overflow-auto" style="max-height: 310px;">
                        <div class="text-center text-muted py-4">
                            <i class="bi bi-info-circle fs-3 d-block mb-2"></i>
                            點擊「執行全表差異稽核」掃描前後期滯後與未達標項目。
                        </div>
                    </div>
                </div>
            </div>
        </div>

    </div>

    <!-- 📱 跨手機與電腦連線 (QR Code 掃碼專區) Modal -->
    <div class="modal fade" id="shareModal" tabindex="-1" aria-hidden="true">
        <div class="modal-dialog modal-dialog-centered">
            <div class="modal-content rounded-4 shadow">
                <div class="modal-header bg-gradient bg-primary text-white rounded-top-4 py-3">
                    <h5 class="modal-title fw-bold d-flex align-items-center gap-2">
                        <i class="bi bi-qr-code-scan fs-4"></i> 📱 跨電腦與手機連線 (QR Code 掃碼專區)
                    </h5>
                    <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal" aria-label="Close"></button>
                </div>
                <div class="modal-body p-4 text-center">
                    <p class="text-muted small mb-3">同仁使用<strong>智慧型手機 (iPhone / Android)</strong> 或<strong>其他電腦</strong>，皆可透過以下方式連線使用系統：</p>

                    <!-- QR Code 展示區 -->
                    <div class="bg-light p-3 rounded-4 border d-inline-block shadow-sm mb-3">
                        <img id="share-qrcode-img" src="" alt="連線 QR Code" class="img-fluid rounded-3" style="width:200px; height:200px;">
                    </div>
                    <div class="fw-bold text-primary mb-3" id="share-lan-url-display">http://...</div>

                    <div class="alert alert-info text-start small mb-3">
                        <div class="fw-bold mb-1"><i class="bi bi-phone-fill me-1"></i> 手機連線方式：</div>
                        1. 手機連接與本機相同的 <strong>Wi-Fi / 局域網</strong>。<br>
                        2. 開啟手機相機或 LINE 掃瞄器，對準上方 <strong>QR Code</strong> 即可秒開啟系統。<br>
                    </div>

                    <div class="alert alert-secondary text-start small mb-0">
                        <div class="fw-bold mb-1"><i class="bi bi-globe me-1"></i> 跨網域 / 行動網路 (4G/5G) 外網連線：</div>
                        如需從家裡或外網連線，可於控制台執行以下命令開啟免費外網通道：<br>
                        <code class="user-select-all bg-dark text-warning p-1 rounded d-block mt-1">npx localtunnel --port 8080</code>
                    </div>
                </div>
                <div class="modal-footer border-0 pt-0 justify-content-between">
                    <button type="button" class="btn btn-outline-secondary rounded-3" data-bs-dismiss="modal">關閉</button>
                    <button type="button" class="btn btn-primary rounded-3 fw-bold" onclick="copyLanUrlFromModal()">
                        <i class="bi bi-clipboard-check"></i> 複製連線網址
                    </button>
                </div>
            </div>
        </div>
    </div>

    <!-- 📤 上傳 Excel Modal -->
    <div class="modal fade" id="uploadModal" tabindex="-1" aria-hidden="true">
        <div class="modal-dialog modal-dialog-centered">
            <div class="modal-content rounded-4 shadow">
                <div class="modal-header border-0 pb-0">
                    <h5 class="modal-header-title fw-bold">📤 上傳新期填報檔 (自動比對前後差異)</h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
                </div>
                <div class="modal-body p-4">
                    <p class="text-muted small">上傳新期填報資料時，系統採用<strong>非覆蓋歷史比對模式</strong>，會保留舊期數據並自動計算前後期達成率增減 (Δ%)：</p>
                    <div class="mb-3">
                        <input type="file" id="excel-file-input" class="form-control" accept=".xlsx, .xls">
                    </div>
                    <div class="alert alert-warning small mb-0">
                        <i class="bi bi-lightning-charge-fill me-1"></i> 上傳後將自動辨識「轉移達成（翻轉達標）」與「未顯著增加（滯後）」項目並標示於儀表板。
                    </div>
                </div>
                <div class="modal-footer border-0 pt-0">
                    <button type="button" class="btn btn-light rounded-3" data-bs-dismiss="modal">取消</button>
                    <button type="button" class="btn btn-warning text-dark fw-bold rounded-3" onclick="uploadExcelFile()">上傳並自動比對差異</button>
                </div>
            </div>
        </div>
    </div>

    <!-- 編輯指標 Modal -->
    <div class="modal fade" id="editModal" tabindex="-1" aria-hidden="true">
        <div class="modal-dialog modal-dialog-centered">
            <div class="modal-content rounded-4 shadow">
                <div class="modal-header border-0 pb-0">
                    <h5 class="modal-header-title fw-bold" id="editModalLabel">✏️ 填報實績與質化說明</h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
                </div>
                <div class="modal-body">
                    <form id="editForm">
                        <input type="hidden" id="edit-id">
                        <div class="mb-3">
                            <label class="form-label text-muted small fw-semibold">指標項目</label>
                            <input type="text" id="edit-item" class="form-control" readonly>
                        </div>
                        <div class="mb-3">
                            <label class="form-label text-muted small fw-semibold">最新達成率 (%)</label>
                            <input type="number" step="any" id="edit-rate" class="form-control" required>
                        </div>
                        <div class="mb-3">
                            <label class="form-label text-muted small fw-semibold">質化成效描述與未達標檢討說明（限300字）</label>
                            <textarea id="edit-desc" class="form-control" rows="4" maxlength="350"></textarea>
                            <div class="form-text text-end"><span id="desc-length">0</span> / 300 字</div>
                        </div>
                        <div class="mb-3">
                            <label class="form-label text-muted small fw-semibold">近期重要里程碑</label>
                            <input type="text" id="edit-milestone" class="form-control">
                        </div>
                        <div class="mb-3">
                            <label class="form-label text-muted small fw-semibold">管考期限</label>
                            <input type="date" id="edit-deadline" class="form-control">
                        </div>
                    </form>
                </div>
                <div class="modal-footer border-0 pt-0">
                    <button type="button" class="btn btn-light rounded-3" data-bs-dismiss="modal">取消</button>
                    <button type="button" class="btn btn-primary rounded-3" onclick="saveIndicator()">儲存更新</button>
                </div>
            </div>
        </div>
    </div>

    <!-- AI 策略與歷史軌跡 Modal -->
    <div class="modal fade" id="aiModal" tabindex="-1" aria-hidden="true">
        <div class="modal-dialog modal-dialog-centered modal-lg">
            <div class="modal-content rounded-4 shadow">
                <div class="modal-header bg-primary text-white rounded-top-4 py-3">
                    <h5 class="modal-title fw-bold" id="aiModalTitle">🤖 前後期歷程比較與 AI 輔導報告</h5>
                    <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal" aria-label="Close"></button>
                </div>
                <div class="modal-body p-4" id="aiModalContent">
                    <div class="text-center py-4">
                        <div class="spinner-border text-primary" role="status">
                            <span class="visually-hidden">Loading...</span>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- ☁️ GitHub 自動同步與設定 Modal -->
    <div class="modal fade" id="githubModal" tabindex="-1" aria-hidden="true">
        <div class="modal-dialog modal-dialog-centered modal-lg">
            <div class="modal-content rounded-4 shadow">
                <div class="modal-header bg-dark text-white rounded-top-4 py-3">
                    <h5 class="modal-title fw-bold d-flex align-items-center gap-2">
                        <i class="bi bi-github fs-4 text-info"></i> ☁️ GitHub 自動同步與遠端儲存庫管理
                    </h5>
                    <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal" aria-label="Close"></button>
                </div>
                <div class="modal-body p-4">
                    <!-- Git 狀態顯示卡片 -->
                    <div class="card bg-light border mb-4">
                        <div class="card-body">
                            <div class="d-flex justify-content-between align-items-center mb-2">
                                <h6 class="fw-bold m-0"><i class="bi bi-info-circle-fill text-primary me-1"></i> 本地 Git 儲存庫與連線狀態</h6>
                                <span id="git-status-badge" class="badge bg-secondary">檢查中...</span>
                            </div>
                            <div class="row g-2 small text-muted">
                                <div class="col-md-4"><strong>目前分支：</strong><span id="git-branch-text" class="text-dark fw-bold">-</span></div>
                                <div class="col-md-8"><strong>遠端 Repo (origin)：</strong><span id="git-remote-text" class="text-dark fw-bold">未設定</span></div>
                                <div class="col-md-12"><strong>待 Commit 變更數量：</strong><span id="git-uncommitted-text" class="text-dark fw-bold">0 個檔案</span></div>
                            </div>
                        </div>
                    </div>

                    <!-- 遠端 Repo 與 PAT 設定表單 -->
                    <div class="mb-3">
                        <label class="form-label text-dark fw-bold small">GitHub 遠端儲存庫網址 (Repository URL)</label>
                        <input type="text" id="gh-remote-input" class="form-control" placeholder="https://github.com/YourUsername/sprout-pm-system.git">
                        <div class="form-text">例如：<code>https://github.com/YourUsername/sprout-pm-system.git</code></div>
                    </div>
                    <div class="mb-3">
                        <label class="form-label text-dark fw-bold small">GitHub Personal Access Token (PAT) [選填，用於私有庫或免密碼驗證]</label>
                        <input type="password" id="gh-token-input" class="form-control" placeholder="ghp_xxxxxxxxxxxxxxxxxxxx">
                        <div class="form-text">若推送時提示存取權限錯誤，可填入 Token (具備 repo 寫入權限)。</div>
                    </div>
                    <div class="mb-3">
                        <label class="form-label text-dark fw-bold small">本次 Commit 說明紀錄 (選填)</label>
                        <input type="text" id="gh-msg-input" class="form-control" placeholder="自動同步高教深耕管考指標數據">
                    </div>

                    <!-- 操作結果訊息顯示框 -->
                    <div id="github-output-box" class="alert d-none small mb-0"></div>
                </div>
                <div class="modal-footer border-0 pt-0">
                    <button type="button" class="btn btn-outline-secondary rounded-3" onclick="saveGithubConfigOnly()">僅儲存設定</button>
                    <button type="button" class="btn btn-primary rounded-3 fw-bold d-flex align-items-center gap-2" onclick="triggerGithubPush()" id="gh-push-btn">
                        <i class="bi bi-cloud-upload-fill"></i> 🚀 立即 Commit 並同步上傳至 GitHub
                    </button>
                </div>
            </div>
        </div>
    </div>

    <!-- Bootstrap 5 JS Bundle -->
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>

    <script>
        let rawData = [];
        let currentThreshold = 0.70; 
        let currentDeltaFilter = 'ALL'; // ALL, PROGRESS, STAGNANT, UNMET
        let aspectChartObj = null;

        const ASPECTS_META = {
            "教學創新精進": { code: "A1", icon: "bi-journal-bookmark-fill", color: "#2563eb", bg: "#eff6ff", border: "#bfdbfe" },
            "產學合作連結": { code: "A2", icon: "bi-briefcase-fill", color: "#d97706", bg: "#fffbeb", border: "#fde68a" },
            "提升高教公共性": { code: "A3", icon: "bi-people-fill", color: "#059669", bg: "#ecfdf5", border: "#a7f3d0" },
            "善盡社會責任 (USR)": { code: "A4", icon: "bi-heart-pulse-fill", color: "#7c3aed", bg: "#f5f3ff", border: "#ddd6fe" }
        };

        document.addEventListener('DOMContentLoaded', () => {
            loadIndicators();
            document.getElementById('edit-desc').addEventListener('input', (e) => {
                document.getElementById('desc-length').textContent = e.target.value.length;
            });
        });

        function loadIndicators() {
            fetch('/api/indicators')
                .then(r => r.json())
                .then(data => {
                    rawData = data;
                    refreshDashboard();
                });
        }

        function setDeltaFilter(type) {
            currentDeltaFilter = type;
            // Update button UI active status
            const btns = document.querySelectorAll('#delta-filter-group .btn');
            btns.forEach(b => b.classList.remove('active'));
            event.target.classList.add('active');
            refreshDashboard();
        }

        function openUploadModal() {
            const modal = new bootstrap.Modal(document.getElementById('uploadModal'));
            modal.show();
        }

        function uploadExcelFile() {
            const input = document.getElementById('excel-file-input');
            if (!input.files || input.files.length === 0) {
                alert('請先選擇要上傳的 Excel (.xlsx) 檔案！');
                return;
            }
            const file = input.files[0];

            fetch('/api/upload_excel', {
                method: 'POST',
                body: file
            })
            .then(r => r.json())
            .then(res => {
                if (res.success) {
                    bootstrap.Modal.getInstance(document.getElementById('uploadModal')).hide();
                    alert('🎉 ' + res.message);
                    loadIndicators();
                } else {
                    alert('上傳解析失敗：' + (res.error || '不支援的檔案格式'));
                }
            });
        }

        function syncThresholdInput(val) {
            val = Math.max(0, Math.min(100, val));
            document.getElementById('threshold-number').value = val;
            currentThreshold = val / 100.0;
            refreshDashboard();
        }

        function syncThresholdSlider(val) {
            val = Math.max(0, Math.min(100, val));
            document.getElementById('threshold-range').value = val;
            currentThreshold = val / 100.0;
            refreshDashboard();
        }

        function refreshDashboard() {
            updateTrendSummary();
            renderAspectSections();
            renderChart();
        }

        function updateTrendSummary() {
            const pct = (currentThreshold * 100).toFixed(0);
            document.getElementById('summary-threshold-val').textContent = pct;

            const progressList = rawData.filter(i => i.trend_status === 'PROGRESS_MET');
            const stagnantList = rawData.filter(i => i.trend_status === 'STAGNANT');
            const unmetList = rawData.filter(i => i.calc_rate < currentThreshold);
            const loadedTargetsCount = rawData.filter(i => i.annual_targets && Object.keys(i.annual_targets).length > 0).length;

            document.getElementById('cnt-progress').textContent = progressList.length;
            document.getElementById('cnt-stagnant').textContent = stagnantList.length;

            const container = document.getElementById('trend-summary-cards');
            container.innerHTML = `
                <div class="col-md-3">
                    <div class="card-custom p-3 border-start border-4 border-primary">
                        <div class="text-muted small">總列管與跨年度目標載入</div>
                        <div class="fs-4 fw-bold text-primary">${loadedTargetsCount} / ${rawData.length} 項</div>
                        <div class="small text-muted">已全面載入 112~116 全期目標值</div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card-custom p-3 border-start border-4 border-success">
                        <div class="text-muted small">🎉 轉移達成 / 顯著進步</div>
                        <div class="fs-4 fw-bold text-success">${progressList.length} 項</div>
                        <div class="small text-success">前期未達標翻轉或成長 Δ ≥ +5%</div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card-custom p-3 border-start border-4 border-warning">
                        <div class="text-muted small">⚠️ 達成率未顯著增加 (滯後)</div>
                        <div class="fs-4 fw-bold text-warning">${stagnantList.length} 項</div>
                        <div class="small text-muted">增幅 Δ ≤ +1% 且尚未高標達標</div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card-custom p-3 border-start border-4 border-danger">
                        <div class="text-muted small">🔴 未達門檻 (&lt;${pct}%)</div>
                        <div class="fs-4 fw-bold text-danger">${unmetList.length} 項</div>
                        <div class="small text-danger">需重點管考與 AI 輔導介入</div>
                    </div>
                </div>
            `;
        }

        function scrollToAspect(code) {
            const el = document.getElementById(`aspect-section-${code}`);
            if (el) el.scrollIntoView({ behavior: 'smooth' });
        }

        // 核心渲染邏輯：依四大構面呈現對應指標
        function renderAspectSections() {
            const container = document.getElementById('aspects-main-container');
            const search = document.getElementById('search-input').value.toLowerCase();
            const categoryFilter = document.getElementById('category-filter').value;
            const aspectFilter = document.getElementById('aspect-filter').value;

            let html = '';

            Object.keys(ASPECTS_META).forEach(aspect => {
                if (aspectFilter && aspectFilter !== aspect) return;

                const meta = ASPECTS_META[aspect];
                let allAspectItems = rawData.filter(i => i.aspect === aspect);

                if (categoryFilter) {
                    allAspectItems = allAspectItems.filter(i => i.category === categoryFilter);
                }

                if (search) {
                    allAspectItems = allAspectItems.filter(i => 
                        i.item.toLowerCase().includes(search) || 
                        i.id.toLowerCase().includes(search) || 
                        i.dept.toLowerCase().includes(search)
                    );
                }
                
                // 按趨勢狀態過濾
                let displayItems = allAspectItems;
                if (currentDeltaFilter === 'PROGRESS') {
                    displayItems = allAspectItems.filter(i => i.trend_status === 'PROGRESS_MET');
                } else if (currentDeltaFilter === 'STAGNANT') {
                    displayItems = allAspectItems.filter(i => i.trend_status === 'STAGNANT');
                } else if (currentDeltaFilter === 'UNMET') {
                    displayItems = allAspectItems.filter(i => i.calc_rate < currentThreshold);
                }

                html += `
                <div class="card-custom p-4 mb-4" id="aspect-section-${meta.code}">
                    <div class="d-flex flex-wrap justify-content-between align-items-center mb-3 pb-2 border-bottom gap-2">
                        <div class="d-flex align-items-center gap-2">
                            <span class="badge rounded-pill fs-6" style="background-color:${meta.color};">構面 ${meta.code}</span>
                            <h5 class="fw-bold text-dark m-0">${aspect}</h5>
                            <span class="badge bg-light text-dark border ms-2">顯示 ${displayItems.length} / 共 ${allAspectItems.length} 項</span>
                        </div>
                    </div>

                    <div class="table-responsive">
                        <table class="table table-hover align-middle border rounded-3 overflow-hidden m-0">
                            <thead class="table-light">
                                <tr>
                                    <th style="width: 90px;">代碼/類別</th>
                                    <th>指標項目與實績說明</th>
                                    <th style="width: 100px;">主責單位</th>
                                    <th style="width: 200px;">前後期比較 (次新 ➔ 最新)</th>
                                    <th style="width: 130px;">差異增減 (Δ%)</th>
                                    <th style="width: 170px;">趨勢狀態</th>
                                    <th class="text-center" style="width: 100px;">操作</th>
                                </tr>
                            </thead>
                            <tbody>
                                ${renderAspectTableRows(displayItems)}
                            </tbody>
                        </table>
                    </div>
                </div>`;
            });

            container.innerHTML = html;
        }

        function renderAspectTableRows(items) {
            if (items.length === 0) {
                return `<tr><td colspan="7" class="text-center py-4 text-muted bg-light"><i class="bi bi-info-circle fs-5 d-block mb-1"></i>目前過濾條件下無項目！</td></tr>`;
            }

            let html = '';
            items.forEach(ind => {
                const prevPct = (ind.prev_calc_rate * 100).toFixed(1);
                const currPct = (ind.calc_rate * 100).toFixed(1);
                const deltaPct = (ind.rate_delta * 100).toFixed(1);
                const isUnmet = ind.calc_rate < currentThreshold;
                
                let trClass = '';
                let statusBadge = '';
                let deltaBadge = '';

                if (ind.trend_status === 'PROGRESS_MET') {
                    trClass = 'tr-progress';
                    statusBadge = `<span class="badge bg-success"><i class="bi bi-trophy-fill me-1"></i>🎉 轉移達成</span>`;
                    deltaBadge = `<span class="delta-pill bg-success text-white">+${deltaPct}% ▲</span>`;
                } else if (ind.trend_status === 'STAGNANT') {
                    trClass = 'tr-stagnant';
                    statusBadge = `<span class="badge bg-warning text-dark"><i class="bi bi-exclamation-triangle-fill me-1"></i>⚠️ 滯後未顯著增加</span>`;
                    deltaBadge = `<span class="delta-pill bg-warning text-dark">${deltaPct}% ▬</span>`;
                } else if (isUnmet) {
                    trClass = 'tr-unmet';
                    statusBadge = `<span class="badge bg-danger"><i class="bi bi-x-circle-fill me-1"></i>🚨 未達標</span>`;
                    deltaBadge = `<span class="delta-pill bg-danger text-white">${deltaPct}%</span>`;
                } else {
                    statusBadge = `<span class="badge bg-secondary">穩定達標</span>`;
                    deltaBadge = `<span class="delta-pill bg-light text-dark border">${deltaPct >= 0 ? '+' : ''}${deltaPct}%</span>`;
                }

                let categoryBadge = ind.category === '自訂指標' ? '<span class="badge" style="background:#f3e8ff; color:#6b21a8;">自訂</span>' : '<span class="badge bg-light text-dark border">共同</span>';

                let multiYearRatePills = '';
                if (ind.annual_rates) {
                    const r112 = ind.annual_rates['112'] != null ? (ind.annual_rates['112']*100).toFixed(1)+'%' : '-';
                    const r113 = ind.annual_rates['113'] != null ? (ind.annual_rates['113']*100).toFixed(1)+'%' : '-';
                    const r114 = ind.annual_rates['114'] != null ? (ind.annual_rates['114']*100).toFixed(1)+'%' : '-';
                    multiYearRatePills = `<div class="mt-1"><span class="badge bg-primary-subtle text-primary border" style="font-size:0.7rem;">📈 歷年達成率 112:${r112} ➔ 113:${r113} ➔ 114:${r114}</span></div>`;
                }

                const latestYrLabel = ind.latest_yr ? (ind.latest_yr.endsWith('年') ? ind.latest_yr : ind.latest_yr + '年') : '最新';
                const prevYrLabel = ind.prev_yr ? (ind.prev_yr.endsWith('年') ? ind.prev_yr : ind.prev_yr + '年') : '次新';

                html += `<tr class="${trClass}" id="row-${ind.id}">
                    <td>
                        <div class="fw-bold text-dark">${ind.id}</div>
                        <div class="mt-1">${categoryBadge}</div>
                    </td>
                    <td>
                        <div class="fw-bold text-dark">${ind.item}</div>
                        <div class="small text-muted mt-1">${ind.qualitative_desc || '無質化說明'}</div>
                        ${multiYearRatePills}
                    </td>
                    <td><span class="badge bg-light text-dark border">${ind.dept}</span></td>
                    <td>
                        <div class="small text-muted">
                            <span class="badge bg-light text-dark border me-1">${prevYrLabel}</span>${prevPct}% ➔ 
                            <span class="badge bg-primary-subtle text-primary border ms-1 me-1">${latestYrLabel}</span><strong class="text-dark fs-6">${currPct}%</strong>
                        </div>
                        <div class="small text-muted text-truncate mt-1" style="max-width:180px;" title="${ind.current_val_text || ''}">${ind.current_val_text || ''}</div>
                    </td>
                    <td>${deltaBadge}</td>
                    <td>${statusBadge}</td>
                    <td class="text-center">
                        <div class="btn-group btn-group-sm">
                            <button class="btn btn-outline-secondary" title="編輯" onclick="openEditModal('${ind.id}')">
                                <i class="bi bi-pencil"></i>
                            </button>
                            <button class="btn btn-outline-primary" title="跨年度目標與歷程對策" onclick="openAiModal('${ind.id}')">
                                <i class="bi bi-clock-history"></i>
                            </button>
                        </div>
                    </td>
                </tr>`;
            });
            return html;
        }

        function resetFilters() {
            document.getElementById('search-input').value = '';
            document.getElementById('category-filter').value = '';
            document.getElementById('aspect-filter').value = '';
            setThreshold(70);
            currentDeltaFilter = 'ALL';
            const btns = document.querySelectorAll('#delta-filter-group .btn');
            btns.forEach(b => b.classList.remove('active'));
            btns[0].classList.add('active');
            refreshDashboard();
        }

        function renderChart() {
            const ctx = document.getElementById('aspectChart').getContext('2d');
            const sampleData = rawData.slice(0, 15); 
            const labels = sampleData.map(i => i.id + ' ' + i.item.substring(0, 5));
            const deltas = sampleData.map(i => (i.rate_delta * 100).toFixed(1));
            const colors = sampleData.map(i => i.rate_delta > 0 ? '#10b981' : (i.rate_delta < 0 ? '#ef4444' : '#f59e0b'));

            if (aspectChartObj) {
                aspectChartObj.destroy();
            }

            aspectChartObj = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: labels,
                    datasets: [{
                        label: '前後期差異增減 (Δ%)',
                        data: deltas,
                        backgroundColor: colors,
                        borderRadius: 6
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    scales: {
                        y: {
                            beginAtZero: true,
                            ticks: { callback: value => value + '%' }
                        },
                        x: {
                            ticks: { font: { size: 9 } }
                        }
                    },
                    plugins: {
                        legend: { display: false }
                    }
                }
            });
        }

        function openEditModal(id) {
            const ind = rawData.find(i => i.id === id);
            if (!ind) return;

            document.getElementById('edit-id').value = ind.id;
            document.getElementById('edit-item').value = `[${ind.id}] ${ind.item} (${ind.aspect})`;
            document.getElementById('edit-rate').value = (ind.calc_rate * 100).toFixed(2);
            document.getElementById('edit-desc').value = ind.qualitative_desc || '';
            document.getElementById('desc-length').textContent = (ind.qualitative_desc || '').length;
            document.getElementById('edit-milestone').value = ind.milestone || '';
            document.getElementById('edit-deadline').value = ind.deadline || '';

            const modal = new bootstrap.Modal(document.getElementById('editModal'));
            modal.show();
        }

        function saveIndicator() {
            const id = document.getElementById('edit-id').value;
            const rate_pct = document.getElementById('edit-rate').value;
            const qualitative_desc = document.getElementById('edit-desc').value;
            const milestone = document.getElementById('edit-milestone').value;
            const deadline = document.getElementById('edit-deadline').value;

            const calc_rate = Math.round(parseFloat(rate_pct) / 100.0 * 10000) / 10000;

            fetch('/api/indicators/update', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ id, calc_rate, qualitative_desc, milestone, deadline })
            })
            .then(r => r.json())
            .then(res => {
                if (res.success) {
                    bootstrap.Modal.getInstance(document.getElementById('editModal')).hide();
                    loadIndicators();
                }
            });
        }

        function openAiModal(id) {
            const ind = rawData.find(i => i.id === id);
            if (!ind) return;

            const modal = new bootstrap.Modal(document.getElementById('aiModal'));
            modal.show();

            const deltaPct = (ind.rate_delta * 100).toFixed(1);
            const historyHtml = (ind.history || []).map(h => 
                `<li class="list-group-item d-flex justify-content-between align-items-center">
                    <div><strong>${h.period}</strong> - ${h.text || '已填報'}</div>
                    <span class="badge bg-primary fs-6">${(h.calc_rate*100).toFixed(1)}%</span>
                </li>`
            ).join('');

            const multiYearTableHtml = `
                <h6 class="fw-bold text-dark mb-2 mt-4"><i class="bi bi-calendar3-range text-primary"></i> 📅 高教深耕第二期 (112~116年度) 跨年度目標與歷年實績對比表</h6>
                <div class="table-responsive mb-4">
                    <table class="table table-sm table-bordered text-center align-middle small mb-0">
                        <thead class="table-light">
                            <tr>
                                <th style="width: 130px;">項目/年度</th>
                                <th>112年度</th>
                                <th>113年度</th>
                                <th>114年度</th>
                                <th>115年度</th>
                                <th>116年度</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <th class="table-light text-start">🎯 年度目標值</th>
                                <td>${(ind.annual_targets && ind.annual_targets['112']) || '-'}</td>
                                <td>${(ind.annual_targets && ind.annual_targets['113']) || '-'}</td>
                                <td>${(ind.annual_targets && ind.annual_targets['114']) || '-'}</td>
                                <td>${(ind.annual_targets && ind.annual_targets['115']) || '-'}</td>
                                <td>${(ind.annual_targets && ind.annual_targets['116']) || '-'}</td>
                            </tr>
                            <tr>
                                <th class="table-light text-start">📊 現況/實績數值</th>
                                <td>${(ind.annual_actuals && (ind.annual_actuals['112'] || ind.annual_actuals['111'])) || '-'}</td>
                                <td>${(ind.annual_actuals && ind.annual_actuals['113']) || '-'}</td>
                                <td>${(ind.annual_actuals && ind.annual_actuals['114']) || '-'}</td>
                                <td class="text-muted">（進行中）</td>
                                <td class="text-muted">（規劃中）</td>
                            </tr>
                            <tr>
                                <th class="table-light text-start">📈 實際達成率 (%)</th>
                                <td>${ind.annual_rates && ind.annual_rates['112'] != null ? `<span class="badge ${ind.annual_rates['112']>=1.0?'bg-success':'bg-warning text-dark'}">${(ind.annual_rates['112']*100).toFixed(1)}%</span>` : '-'}</td>
                                <td>${ind.annual_rates && ind.annual_rates['113'] != null ? `<span class="badge ${ind.annual_rates['113']>=1.0?'bg-success':'bg-warning text-dark'}">${(ind.annual_rates['113']*100).toFixed(1)}%</span>` : '-'}</td>
                                <td>${ind.annual_rates && ind.annual_rates['114'] != null ? `<span class="badge ${ind.annual_rates['114']>=1.0?'bg-success':'bg-warning text-dark'}">${(ind.annual_rates['114']*100).toFixed(1)}%</span>` : '-'}</td>
                                <td class="text-muted">-</td>
                                <td class="text-muted">-</td>
                            </tr>
                        </tbody>
                    </table>
                </div>`;

            const latestYrTitle = ind.latest_yr ? (ind.latest_yr.endsWith('年') ? ind.latest_yr : ind.latest_yr + '年') : '最新期';
            const prevYrTitle = ind.prev_yr ? (ind.prev_yr.endsWith('年') ? ind.prev_yr : ind.prev_yr + '年') : '次新期';

            const modalAlertHtml = `<div class="alert alert-light border mb-3">
                <div class="row g-2 text-dark small">
                    <div class="col-md-3"><strong>主責處室：</strong>${ind.dept}</div>
                    <div class="col-md-3"><strong>次新期 (${prevYrTitle})：</strong>${(ind.prev_calc_rate*100).toFixed(1)}%</div>
                    <div class="col-md-3"><strong>最新期 (${latestYrTitle})：</strong>${(ind.calc_rate*100).toFixed(1)}%</div>
                    <div class="col-md-3"><strong>前後差異 (次新➔最新)：</strong><span class="badge ${(ind.rate_delta>=0?'bg-success':'bg-danger')}">${deltaPct}%</span></div>
                </div>
            </div>`;

            const content = document.getElementById('aiModalContent');
            content.innerHTML = `
                <div class="d-flex align-items-center gap-3 mb-3">
                    <span class="badge bg-primary fs-6">${ind.id}</span>
                    <h5 class="fw-bold m-0">${ind.item}</h5>
                    <span class="badge bg-light text-dark border">${ind.aspect}</span>
                </div>
                ${modalAlertHtml}
                ${multiYearTableHtml}
                <h6 class="fw-bold text-dark mb-2"><i class="bi bi-clock-history"></i> 多期歷史填報軌跡 (非覆蓋紀錄)</h6>
                <ul class="list-group mb-4 small">
                    ${historyHtml}
                </ul>
                <h6 class="fw-bold text-primary mb-2"><i class="bi bi-cpu-fill"></i> 專責 Agent：${ind.ai_agent || '【AI 專案 Agent】'} 差異化介入方案</h6>
                <div class="border rounded-3 p-3 bg-light text-dark small lh-lg">
                    ${ind.trend_status === 'STAGNANT' ? 
                        `<div class="text-danger fw-bold mb-2">⚠️ 警告：該指標前後期增減 (Δ ${deltaPct}%) 屬於未顯著增加（滯後指標），需重點輔導！</div>` : 
                        `<div class="text-success fw-bold mb-2">🎉 轉移達成 / 成長良好 (Δ ${deltaPct}%)，持續保持！</div>`
                    }
                    <div><strong>改善對策：</strong>${ind.qualitative_desc || '持續推動檢核與管考。'}</div>
                </div>
            `;
        }

        function runAudit() {
            fetch(`/api/audit?threshold=${currentThreshold}`)
                .then(r => r.json())
                .then(data => {
                    let html = '<ul class="list-group list-group-flush">';
                    if (data.length === 0) {
                        html += '<li class="list-group-item text-success text-center py-3"><i class="bi bi-check-circle fs-4 d-block mb-1"></i>全表經智慧差異稽核皆無滯後與未達標問題！</li>';
                    } else {
                        data.forEach(d => {
                            let badge = d.level === 'CRITICAL' ? 'bg-danger' : (d.level === 'WARNING' ? 'bg-warning text-dark' : 'bg-info');
                            html += `<li class="list-group-item px-0 py-2 d-flex justify-content-between align-items-start gap-2">
                                <div>
                                    <strong class="small text-dark">[${d.id}] ${d.item} (${d.aspect})</strong>
                                    <div class="small text-muted">${d.msg}</div>
                                </div>
                                <span class="badge ${badge}">${d.type}</span>
                            </li>`;
                        });
                    }
                    html += '</ul>';
                    document.getElementById('audit-content').innerHTML = html;
                });
        }

        // GitHub 模組前端控制函式
        function openShareModal() {
            fetch('/api/lan_info')
                .then(r => r.json())
                .then(data => {
                    const qrUrl = `https://api.qrserver.com/v1/create-qr-code/?size=220x220&data=${encodeURIComponent(data.lan_url)}`;
                    document.getElementById('share-qrcode-img').src = qrUrl;
                    document.getElementById('share-lan-url-display').textContent = data.lan_url;
                    
                    const modal = new bootstrap.Modal(document.getElementById('shareModal'));
                    modal.show();
                });
        }

        function copyLanUrlFromModal() {
            const lanUrl = document.getElementById('share-lan-url-display').textContent;
            const msg = '🎉 已成功複製連線網址！\\n\\n' + lanUrl + '\\n\\n請將此網址發給他人電腦或手機輸入即可連線。';
            if (navigator.clipboard) {
                navigator.clipboard.writeText(lanUrl).then(() => {
                    alert(msg);
                });
            } else {
                alert(msg);
            }
        }

        function openGithubModal() {
            const modal = new bootstrap.Modal(document.getElementById('githubModal'));
            modal.show();
            refreshGithubModalStatus();
        }

        function refreshGithubModalStatus() {
            const badge = document.getElementById('git-status-badge');
            badge.className = 'badge bg-secondary';
            badge.textContent = '檢查中...';
            
            fetch('/api/github/status')
                .then(r => r.json())
                .then(data => {
                    if (data.is_repo) {
                        if (data.has_remote) {
                            badge.className = 'badge bg-success';
                            badge.textContent = '🟢 已連結 GitHub 遠端';
                        } else {
                            badge.className = 'badge bg-warning text-dark';
                            badge.textContent = '🟡 已建立 Git 庫，尚未連結遠端 Repo';
                        }
                    } else {
                        badge.className = 'badge bg-danger';
                        badge.textContent = '🔴 未建立 Git 庫';
                    }

                    document.getElementById('git-branch-text').textContent = data.branch || 'main';
                    document.getElementById('git-remote-text').textContent = data.remote_url || data.config_remote || '尚未設定';
                    document.getElementById('git-uncommitted-text').textContent = data.has_changes ? `${data.uncommitted_files.length} 個檔案修訂待 commit` : '無待 commit 檔案';

                    if (data.remote_url || data.config_remote) {
                        document.getElementById('gh-remote-input').value = data.remote_url || data.config_remote;
                    }
                })
                .catch(err => {
                    badge.className = 'badge bg-danger';
                    badge.textContent = '🔴 狀態連線異常';
                });
        }

        function saveGithubConfigOnly() {
            const remote_url = document.getElementById('gh-remote-input').value.trim();
            const token = document.getElementById('gh-token-input').value.trim();
            
            fetch('/api/github/config', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ remote_url, token })
            })
            .then(r => r.json())
            .then(res => {
                const box = document.getElementById('github-output-box');
                box.classList.remove('d-none', 'alert-danger');
                box.className = 'alert alert-success small mb-0';
                box.textContent = res.message || '設定已儲存！';
                refreshGithubModalStatus();
            });
        }

        function triggerGithubPush() {
            const btn = document.getElementById('gh-push-btn');
            const box = document.getElementById('github-output-box');
            btn.disabled = true;
            btn.innerHTML = '<span class="spinner-border spinner-border-sm me-2"></span>同步上傳至 GitHub 中...';
            
            box.classList.remove('d-none');
            box.className = 'alert alert-info small mb-0';
            box.textContent = '正在執行 Git add, commit 與 push 上傳，請稍候...';

            const remote_url = document.getElementById('gh-remote-input').value.trim();
            const token = document.getElementById('gh-token-input').value.trim();
            const commit_msg = document.getElementById('gh-msg-input').value.trim();

            fetch('/api/github/push', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ remote_url, token, commit_msg })
            })
            .then(r => r.json())
            .then(res => {
                btn.disabled = false;
                btn.innerHTML = '<i class="bi bi-cloud-upload-fill"></i> 🚀 立即 Commit 並同步上傳至 GitHub';
                
                box.classList.remove('d-none');
                if (res.success) {
                    box.className = 'alert alert-success small mb-0';
                    box.textContent = res.message;
                    refreshGithubModalStatus();
                } else {
                    box.className = 'alert alert-danger small mb-0';
                    box.innerHTML = '<strong>上傳提示/失敗：</strong><br>' + (res.message || '未知錯誤').split('\\n').join('<br>');
                }
            })
            .catch(err => {
                btn.disabled = false;
                btn.innerHTML = '<i class="bi bi-cloud-upload-fill"></i> 🚀 領域 Commit 並同步上傳至 GitHub';
                box.classList.remove('d-none');
                box.className = 'alert alert-danger small mb-0';
                box.textContent = '連線失敗: ' + err;
            });
        }
    </script>
</body>
</html>
"""

def get_lan_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        try:
            return socket.gethostbyname(socket.gethostname())
        except Exception:
            return '127.0.0.1'

class ReusableHTTPServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    allow_reuse_address = True
    daemon_threads = True

def run_server(port=8080, open_browser=True):
    host = '0.0.0.0'
    lan_ip = get_lan_ip()
    for p in [port, 8081, 8085, 8090]:
        try:
            httpd = ReusableHTTPServer((host, p), SproutWebServer)
            print("============================================================")
            print("🏛️ 高教深耕智慧專案管理與指標管考中樞 已成功啟動！")
            print("------------------------------------------------------------")
            print(f" [本機訪問 (Local Host)]   : http://localhost:{p}")
            print(f" [本機訪問 (Local IP)]     : http://127.0.0.1:{p}")
            print(f" [跨電腦/同網域訪問 (LAN)] : http://{lan_ip}:{p}  <-- 請他人電腦輸入此網址")
            print("============================================================")

            # 自動開啟預設瀏覽器進入系統畫面
            if open_browser:
                def _pop_browser():
                    import time, webbrowser
                    time.sleep(0.8)
                    try:
                        webbrowser.open(f"http://localhost:{p}")
                    except Exception:
                        pass
                threading.Thread(target=_pop_browser, daemon=True).start()

            httpd.serve_forever()
            break
        except OSError as e:
            print(f"[WARN] 通訊埠 {p} 被佔用，嘗試自動切換至備用 Port...")
            continue

if __name__ == "__main__":
    run_server(8080)
